#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║                                                                      ║
║   ██╗███╗   ██╗███████╗████████╗  ██████╗ ██████╗ ███████╗██████╗   ║
║   ██║████╗  ██║██╔════╝╚══██╔══╝ ██╔════╝ ██╔══██╗██╔════╝██╔══██╗  ║
║   ██║██╔██╗ ██║███████╗   ██║    ██║  ███╗██████╔╝█████╗  ██████╔╝  ║
║   ██║██║╚██╗██║╚════██║   ██║    ██║   ██║██╔══██╗██╔══╝  ██╔═══╝   ║
║   ██║██║ ╚████║███████║   ██║    ╚██████╔╝██║  ██║███████╗██║       ║
║   ╚═╝╚═╝  ╚═══╝╚══════╝   ╚═╝     ╚═════╝ ╚═╝  ╚═╝╚══════╝╚═╝       ║
║                                                                      ║
║   Instagram OSINT Tool  v4.0  —  Cookie-based authentication        ║
╚══════════════════════════════════════════════════════════════════════╝

Requirements:
    pip install customtkinter instagrapi matplotlib pillow requests
    pip install browser-cookie3 openpyxl reportlab tkinterdnd2

Keyboard shortcuts:
  Enter / Ctrl+Return  — Analyze
  Ctrl+S               — Save JSON
  Ctrl+D               — Download media
  Ctrl+E               — Export PDF
  Ctrl+H               — Toggle recent searches
  ?                    — Shortcut cheatsheet

⚠  Automated access may violate Instagram's Terms of Service.
"""

import io, json, os, queue, re, threading, time, traceback, math
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import requests
from PIL import Image, ImageDraw, ImageFilter, ImageEnhance, ImageTk
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
import numpy as np

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog

# ── Optional dependencies ─────────────────────────────────────────────────────
try:
    from instagrapi import Client
    from instagrapi.exceptions import UserNotFound, LoginRequired, PrivateAccount
    INSTAGRAPI_OK = True
except ImportError:
    INSTAGRAPI_OK = False

try:
    import browser_cookie3
    COOKIES_OK = True
except ImportError:
    COOKIES_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, Image as RLImage, HRFlowable)
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_OK = True
except ImportError:
    DND_OK = False

# ─────────────────────────────────────────────────────────────────────────────
#  Palette & constants
# ─────────────────────────────────────────────────────────────────────────────

APP_TITLE   = "InstaGrep  v4.0"
APP_VERSION = "4.0.0"

C = {
    # backgrounds
    "bg":        "#06060e",
    "bg2":       "#09091a",
    "bg3":       "#0d0d1e",
    "card":      "#111124",
    "card2":     "#171730",
    "card3":     "#1c1c38",
    # borders
    "border":    "#1e1e48",
    "border2":   "#282860",
    "border_hi": "#3838a0",
    # text
    "txt":       "#f0f2ff",
    "txt2":      "#9ba8d8",
    "txt3":      "#525899",
    "txt4":      "#363870",
    # accents
    "cyan":      "#00f5ff",
    "cyan2":     "#00c8d4",
    "cyan_dim":  "#00232a",
    "blue":      "#4d8dff",
    "blue2":     "#2563ff",
    "blue3":     "#142080",
    "green":     "#39ffac",
    "green2":    "#00e676",
    "green_dim": "#002818",
    "purple":    "#c040f5",
    "purple2":   "#8800cc",
    "purple_dim":"#2a0048",
    "pink":      "#ff3d8a",
    "orange":    "#ff9500",
    "yellow":    "#ffd700",
    "red":       "#ff2244",
    "red2":      "#bb0022",
    # status
    "s_ok":      "#39ffac",
    "s_warn":    "#ffd700",
    "s_err":     "#ff2244",
    "s_info":    "#00f5ff",
    # chart
    "ch_bg":     "#06060e",
    "ch_bg2":    "#09091a",
    "ch_axis":   "#1e1e48",
}

# Neon ring colours for avatar
RING_GRAD = ["#00f5ff", "#4d8dff", "#c855f5", "#ff4d8d"]

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
EMAIL_RE  = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b")
PHONE_RE  = re.compile(
    r"(?:\+?1[-.\s]?)?\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}"
    r"|\+[1-9][0-9]{7,14}"
)

BROWSER_LOADERS = []
if COOKIES_OK:
    for _n in ("chrome","firefox","chromium","brave","edge","opera","vivaldi","librewolf"):
        _fn = getattr(browser_cookie3, _n, None)
        if _fn:
            BROWSER_LOADERS.append((_n.title(), _fn))


# ─────────────────────────────────────────────────────────────────────────────
#  Utility helpers
# ─────────────────────────────────────────────────────────────────────────────

def _short(exc, n=140):
    s = str(exc)
    return s[:n] + ("…" if len(s) > n else "")

def fmt_num(n):
    try: n = int(n or 0)
    except: n = 0
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000:     return f"{n/1_000:.1f}K"
    return f"{n:,}"

def time_ago(dt: datetime) -> str:
    if dt is None: return "unknown"
    try:
        now   = datetime.now()
        delta = now - dt.replace(tzinfo=None)
        secs  = int(delta.total_seconds())
        if secs < 60:     return "just now"
        if secs < 3600:   return f"{secs//60}m ago"
        if secs < 86400:  return f"{secs//3600}h ago"
        if secs < 172800: return "yesterday"
        return dt.strftime("%b %d")
    except: return "—"

def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def lerp_color(c1, c2, t):
    r1,g1,b1 = hex_to_rgb(c1)
    r2,g2,b2 = hex_to_rgb(c2)
    return (int(r1+(r2-r1)*t), int(g1+(g2-g1)*t), int(b1+(b2-b1)*t))


def _tb_set(tb, text):
    tb.configure(state="normal")
    tb.delete("1.0", "end")
    tb.insert("1.0", text)
    tb.configure(state="disabled")


def retry(fn, *args, attempts=4, base_delay=1.5, **kwargs):
    last_exc = None
    for i in range(attempts):
        try:
            return fn(*args, **kwargs)
        except (LoginRequired, PrivateAccount):
            raise
        except Exception as e:
            last_exc = e
            if i < attempts-1:
                time.sleep(base_delay * (2**i) + 0.3*i)
    raise last_exc


# ─────────────────────────────────────────────────────────────────────────────
#  Recent searches manager
# ─────────────────────────────────────────────────────────────────────────────

class RecentSearches:
    MAX = 10
    PATH = Path("recent_searches.json")

    def __init__(self):
        self._items: list[dict] = []   # [{username, timestamp, full_name}]
        self._load()

    def _load(self):
        try:
            if self.PATH.exists():
                self._items = json.loads(self.PATH.read_text())
        except: pass

    def _save(self):
        try:
            self.PATH.write_text(json.dumps(self._items, indent=2))
        except: pass

    def add(self, username: str, full_name: str = ""):
        self._items = [x for x in self._items if x.get("username") != username]
        self._items.insert(0, {
            "username":   username,
            "full_name":  full_name,
            "timestamp":  datetime.now().isoformat(),
        })
        self._items = self._items[:self.MAX]
        self._save()

    def get_all(self) -> list[dict]:
        return list(self._items)

    def clear(self):
        self._items = []
        self._save()


# ─────────────────────────────────────────────────────────────────────────────
#  Session manager
# ─────────────────────────────────────────────────────────────────────────────

class SessionManager:
    def __init__(self):
        self._sessions: dict = {}
        self._active: str | None = None
        self._load()

    def _load(self):
        p = Path("instagrep_sessions.json")
        if p.exists():
            try:
                d = json.loads(p.read_text())
                self._sessions = d.get("sessions", {})
                self._active   = d.get("active")
                if self._active not in self._sessions:
                    self._active = next(iter(self._sessions), None)
            except: pass

    def _save(self):
        try:
            Path("instagrep_sessions.json").write_text(json.dumps(
                {"sessions": self._sessions, "active": self._active}, indent=2))
        except: pass

    def add(self, name, cookies):
        self._sessions[name] = cookies
        if self._active is None: self._active = name
        self._save()

    def set_active(self, name):
        if name in self._sessions: self._active = name; self._save()

    def remove(self, name):
        self._sessions.pop(name, None)
        if self._active == name:
            self._active = next(iter(self._sessions), None)
        self._save()

    def get_active(self) -> dict | None:
        return self._sessions.get(self._active)

    def list(self) -> list:
        return list(self._sessions.keys())

    @property
    def active_name(self): return self._active

    def has(self): return bool(self._sessions)


# ─────────────────────────────────────────────────────────────────────────────
#  Download manager
# ─────────────────────────────────────────────────────────────────────────────

class DownloadManager:
    def __init__(self, status_cb=None, progress_cb=None):
        self._status   = status_cb   or (lambda m: None)
        self._progress = progress_cb or (lambda d, t: None)
        self._stop     = False
        self._sess     = requests.Session()
        self._sess.headers.update({"User-Agent": "Mozilla/5.0"})

    def cancel(self): self._stop = True

    def download_file(self, url, dest: Path, timeout=30) -> bool:
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            if dest.exists(): return True
            r = self._sess.get(url, timeout=timeout, stream=True)
            r.raise_for_status()
            with open(dest, "wb") as fh:
                for chunk in r.iter_content(65536):
                    if self._stop: return False
                    fh.write(chunk)
            return True
        except Exception as e:
            self._status(f"✗ {dest.name} — {_short(e,50)}")
            return False

    def download_profile_pic(self, url, username) -> Path | None:
        if not url or url == "None": return None
        ext  = Path(urlparse(url).path).suffix or ".jpg"
        dest = Path(username) / "Profile" / f"profile_pic{ext}"
        return dest if self.download_file(url, dest) else None

    def download_medias(self, medias, username, cl=None) -> dict:
        base = Path(username) / "Posts"
        base.mkdir(parents=True, exist_ok=True)
        total, done, failed = len(medias), 0, 0
        for idx, m in enumerate(medias, 1):
            if self._stop: break
            self._progress(idx, total)
            ta    = getattr(m,"taken_at",None)
            ds    = ta.strftime("%Y%m%d_%H%M%S") if ta else f"post_{idx}"
            mid   = str(getattr(m,"pk",idx))
            mtype = getattr(m,"media_type",1)
            urls  = []
            if mtype == 8:
                for res in (getattr(m,"resources",[]) or []):
                    u = str(getattr(res,"thumbnail_url","") or getattr(res,"video_url","") or "")
                    if u: urls.append(u)
            elif mtype == 2:
                u = str(getattr(m,"video_url","") or "")
                if u: urls.append(u)
            else:
                iv = getattr(m,"image_versions2",None)
                if iv and hasattr(iv,"candidates") and iv.candidates:
                    u = str(iv.candidates[0].url or "")
                    if u and u != "None": urls.append(u)
            for i, url in enumerate(urls):
                ext = ".mp4" if mtype==2 else ".jpg"
                sfx = f"_{i+1}" if len(urls)>1 else ""
                ok  = self.download_file(url, base/f"{ds}_{mid}{sfx}{ext}")
                if ok: done += 1
                else: failed += 1
        return {"total": total, "downloaded": done, "failed": failed}


# ─────────────────────────────────────────────────────────────────────────────
#  Avatar image builder  (neon ring + glow)
# ─────────────────────────────────────────────────────────────────────────────

def make_avatar_ring(img: Image.Image, size=160) -> Image.Image:
    """Paste a circular photo inside a neon gradient ring."""
    sz   = size + 16      # canvas includes ring
    out  = Image.new("RGBA", (sz, sz), (0, 0, 0, 0))
    draw = ImageDraw.Draw(out)

    # Outer glow (soft cyan blur)
    glow = Image.new("RGBA", (sz, sz), (0, 0, 0, 0))
    gd   = ImageDraw.Draw(glow)
    for r in range(8, 0, -1):
        alpha = int(40 * (1 - r/8))
        gd.ellipse([sz//2-size//2-r, sz//2-size//2-r,
                    sz//2+size//2+r, sz//2+size//2+r],
                   outline=(0, 245, 255, alpha), width=1)
    glow = glow.filter(ImageFilter.GaussianBlur(3))
    out.paste(glow, (0, 0), glow)

    # Neon ring (gradient segments)
    n_seg = 120
    cx, cy = sz//2, sz//2
    r_out  = size//2 + 6
    r_in   = size//2 + 2
    ring_colors = [(0,245,255),(77,141,255),(200,85,245),(255,77,141)]
    for i in range(n_seg):
        t     = i / n_seg
        idx   = t * (len(ring_colors)-1)
        lo    = int(idx); hi = min(lo+1, len(ring_colors)-1)
        ft    = idx - lo
        c1, c2 = ring_colors[lo], ring_colors[hi]
        col   = tuple(int(c1[k]+(c2[k]-c1[k])*ft) for k in range(3)) + (255,)
        a1    = math.radians(360*i/n_seg - 90)
        a2    = math.radians(360*(i+1)/n_seg - 90)
        x1,y1 = cx+r_out*math.cos(a1), cy+r_out*math.sin(a1)
        x2,y2 = cx+r_out*math.cos(a2), cy+r_out*math.sin(a2)
        x3,y3 = cx+r_in*math.cos(a2),  cy+r_in*math.sin(a2)
        x4,y4 = cx+r_in*math.cos(a1),  cy+r_in*math.sin(a1)
        draw.polygon([(x1,y1),(x2,y2),(x3,y3),(x4,y4)], fill=col)

    # Circular photo
    photo = img.convert("RGBA").resize((size, size), Image.LANCZOS)
    mask  = Image.new("L", (size, size), 0)
    ImageDraw.Draw(mask).ellipse([0, 0, size-1, size-1], fill=255)
    photo.putalpha(mask)
    off   = sz//2 - size//2
    out.paste(photo, (off, off), photo)
    return out


def make_placeholder_avatar(size=160) -> Image.Image:
    """Simple cyber-styled placeholder avatar."""
    img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.ellipse([0, 0, size-1, size-1], fill=(20, 20, 42))
    cx, cy = size//2, size//2
    # head
    draw.ellipse([cx-24,cy-42,cx+24,cy-8],  fill=(60,60,110))
    # body
    draw.ellipse([cx-38,cy+6,cx+38,cy+76],  fill=(60,60,110))
    mask = Image.new("L",(size,size),0)
    ImageDraw.Draw(mask).ellipse([0,0,size-1,size-1],fill=255)
    img.putalpha(mask)
    return img


# ─────────────────────────────────────────────────────────────────────────────
#  Image utilities  (gradient hero, shimmer, initials avatar, glass card)
# ─────────────────────────────────────────────────────────────────────────────

def make_hero_banner(w=900, h=180) -> Image.Image:
    """Radial + linear gradient hero banner for the profile header."""
    img  = Image.new("RGBA", (w, h), (6, 6, 14, 255))
    draw = ImageDraw.Draw(img)
    # Left cyan glow
    for r in range(200, 0, -4):
        alpha = int(55 * (1 - r/200))
        bbox  = [0-r, h//2-r, r, h//2+r]
        draw.ellipse(bbox, fill=(0, 245, 255, alpha))
    # Right purple glow
    for r in range(220, 0, -4):
        alpha = int(45 * (1 - r/220))
        bbox  = [w-r, -r, w+r, h+r]
        draw.ellipse(bbox, fill=(192, 64, 245, alpha))
    # Bottom-center blue glow
    for r in range(160, 0, -4):
        alpha = int(35 * (1 - r/160))
        bbox  = [w//2-r, h-r//2, w//2+r, h+r//2]
        draw.ellipse(bbox, fill=(77, 141, 255, alpha))
    # Horizontal scan line grid
    for y in range(0, h, 20):
        draw.line([(0,y),(w,y)], fill=(255,255,255,6), width=1)
    img = img.filter(ImageFilter.GaussianBlur(4))
    return img


def make_glass_card_bg(w=300, h=120, accent_hex="#00f5ff") -> Image.Image:
    """Semi-transparent glass-effect card background."""
    img  = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    r1, g1, b1 = hex_to_rgb(accent_hex)
    # Base fill
    draw.rounded_rectangle([0, 0, w-1, h-1], radius=14,
                            fill=(r1//8+10, g1//8+10, b1//8+18, 220))
    # Top highlight strip
    draw.rounded_rectangle([0, 0, w-1, 3], radius=0,
                            fill=(r1, g1, b1, 90))
    # Subtle inner glow
    for depth in range(1, 6):
        a = int(30 * (1 - depth/6))
        draw.rounded_rectangle([depth, depth, w-1-depth, h-1-depth],
                                radius=max(14-depth,0),
                                outline=(r1, g1, b1, a), width=1)
    return img


def make_shimmer_bar(w=260, h=16, radius=6) -> Image.Image:
    """Single animated shimmer bar (placeholder while loading)."""
    img  = Image.new("RGBA", (w, h))
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle([0, 0, w-1, h-1], radius=radius,
                            fill=(30, 30, 60, 255))
    # Sweep highlight
    for i in range(w//3):
        t    = i / (w//3)
        alpha = int(120 * math.sin(math.pi * t))
        draw.line([(i + w//4, 0), (i + w//4, h)],
                  fill=(180, 200, 255, alpha), width=1)
    return img


def make_initials_avatar(letter: str, size=36, hue_seed=0) -> Image.Image:
    """Coloured circle with an initial letter — for recent search list."""
    palettes = [
        (0,245,255),(77,141,255),(192,64,245),(255,61,138),
        (57,255,172),(255,149,0),(255,210,0),(0,230,118),
    ]
    col = palettes[hue_seed % len(palettes)]
    img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.ellipse([0, 0, size-1, size-1], fill=col + (255,))
    # Darken slightly
    ov = Image.new("RGBA", (size, size), (0, 0, 0, 100))
    mask = Image.new("L", (size, size), 0)
    ImageDraw.Draw(mask).ellipse([0, 0, size-1, size-1], fill=255)
    img.paste(ov, (0, 0), mask)
    # Letter
    try:
        from PIL import ImageFont
        fnt = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                                 size=size//2)
    except Exception:
        fnt = None
    letter = (letter or "?")[0].upper()
    if fnt:
        bb  = draw.textbbox((0, 0), letter, font=fnt)
        tx  = (size - (bb[2]-bb[0])) // 2 - bb[0]
        ty  = (size - (bb[3]-bb[1])) // 2 - bb[1]
        draw.text((tx, ty), letter, font=fnt, fill=(255, 255, 255, 240))
    else:
        draw.text((size//4, size//6), letter, fill=(255,255,255,240))
    return img


def make_sparkline(values: list, w=90, h=30,
                   color="#00f5ff") -> Image.Image:
    """Tiny sparkline chart as a PIL image."""
    img  = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    if not values or max(values) == 0:
        return img
    draw = ImageDraw.Draw(img)
    mn   = min(values); mx = max(values)
    span = mx - mn or 1
    pts  = []
    for i, v in enumerate(values):
        x = int(i / max(len(values)-1, 1) * (w-2)) + 1
        y = int((1 - (v - mn) / span) * (h - 4)) + 2
        pts.append((x, y))
    r, g, b = hex_to_rgb(color)
    # Glow fill under line
    for x, y in pts:
        for yy in range(y, h):
            alpha = max(0, int(60 * (1 - (yy-y)/(h-y+1))))
            px = img.getpixel((x, yy))
            img.putpixel((x, yy), (r, g, b, min(px[3]+alpha, 180)))
    # Line
    if len(pts) >= 2:
        for i in range(len(pts)-1):
            draw.line([pts[i], pts[i+1]], fill=(r,g,b,230), width=2)
    # Endpoint dot
    if pts:
        ex, ey = pts[-1]
        draw.ellipse([ex-2,ey-2,ex+2,ey+2], fill=(r,g,b,255))
    return img


# ─────────────────────────────────────────────────────────────────────────────
#  Chart engine  (v4 — professional quality)
# ─────────────────────────────────────────────────────────────────────────────

class ChartEngine:

    # ── Shared style ──────────────────────────────────────────────────────────
    @staticmethod
    def _setup_ax(ax, title, title_color):
        ax.set_facecolor(C["ch_bg2"])
        for spine in ("top","right"):
            ax.spines[spine].set_visible(False)
        for spine in ("bottom","left"):
            ax.spines[spine].set_color(C["ch_axis"])
            ax.spines[spine].set_linewidth(0.6)
        ax.tick_params(colors=C["txt3"], length=3, width=0.5, labelsize=8, pad=4)
        ax.yaxis.grid(True, color=C["ch_axis"], linewidth=0.3,
                      linestyle="--", alpha=0.8, zorder=0)
        ax.set_axisbelow(True)
        ax.set_title(title, color=title_color, fontsize=12,
                     fontweight="bold", pad=12, loc="left",
                     fontfamily="monospace")

    @staticmethod
    def _base_fig(w=10.0, h=3.2):
        fig, ax = plt.subplots(figsize=(w, h))
        fig.patch.set_facecolor(C["ch_bg"])
        fig.subplots_adjust(left=0.06, right=0.97, top=0.85, bottom=0.18)
        return fig, ax

    @staticmethod
    def _save(fig) -> Image.Image:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130,
                    facecolor=C["ch_bg"], edgecolor="none")
        buf.seek(0)
        img = Image.open(buf).copy()
        buf.close()
        plt.close(fig)
        return img

    @classmethod
    def _gradient_bars(cls, ax, xs, heights, base_hex, peak_hex,
                       width=0.75, peak_idx=None):
        """Draw bars with per-segment vertical gradient via imshow trick."""
        r1,g1,b1 = hex_to_rgb(base_hex)
        r2,g2,b2 = hex_to_rgb(peak_hex)
        mx = max(heights) if heights else 1
        for i, (x, h) in enumerate(zip(xs, heights)):
            if h <= 0: continue
            is_peak = (peak_idx is not None and i == peak_idx)
            col     = peak_hex if is_peak else base_hex
            # Gradient image for this bar
            grad = np.zeros((100, 1, 4))
            for row in range(100):
                t = row / 100
                if is_peak:
                    rc = int(r1*(1-t)+r2*t)
                    gc = int(g1*(1-t)+g2*t)
                    bc = int(b1*(1-t)+b2*t)
                    alpha = 0.95
                else:
                    rc = max(r1//3, int(r2*t*0.7))
                    gc = max(g1//3, int(g2*t*0.7))
                    bc = max(b1//3, int(b2*t*0.7))
                    alpha = 0.80
                grad[99-row, 0] = [rc/255, gc/255, bc/255, alpha]
            ax.imshow(grad, extent=[x-width/2, x+width/2, 0, h],
                      aspect="auto", zorder=3, interpolation="bilinear")
            # Outer glow on peak
            if is_peak:
                ax.bar(x, h, width=width+0.25, color=peak_hex,
                       alpha=0.08, zorder=2, edgecolor="none")

    @classmethod
    def hours_chart(cls, hours: list) -> Image.Image:
        fig, ax = cls._base_fig(10.0, 3.4)
        mx = max(hours) if any(hours) else 1
        pi = hours.index(mx) if mx > 0 else -1

        cls._gradient_bars(ax, range(24), hours,
                           "#003844", "#00f5ff", width=0.72, peak_idx=pi)

        # Average line
        avg = sum(hours) / max(sum(1 for h in hours if h>0), 1)
        ax.axhline(avg, color="#4d8dff", linewidth=0.9, linestyle=":",
                   alpha=0.7, zorder=1)
        ax.text(23.6, avg + mx*0.03, f"avg {avg:.1f}",
                ha="right", va="bottom", fontsize=7,
                color="#4d8dff", fontstyle="italic")

        # Value labels
        for i, v in enumerate(hours):
            if v > mx * 0.15:
                ax.text(i, v + mx*0.025, str(v), ha="center", va="bottom",
                        fontsize=6.5,
                        color="#00f5ff" if i==pi else C["txt3"],
                        fontweight="bold" if i==pi else "normal")

        # Peak annotation
        if pi >= 0:
            ax.annotate(f"Peak  {pi:02d}:00",
                        xy=(pi, mx),
                        xytext=(pi+2.5 if pi<18 else pi-2.5, mx*0.88),
                        fontsize=8, color="#00f5ff", fontweight="bold",
                        arrowprops=dict(arrowstyle="-|>", color="#00f5ff",
                                        lw=0.8, mutation_scale=10))

        ax.set_xticks(range(24))
        ax.set_xticklabels([f"{h:02d}" for h in range(24)],
                           color=C["txt3"], fontsize=7.5)
        ax.set_xlim(-0.7, 23.7)
        ax.set_ylim(0, mx * 1.22)
        ax.set_ylabel("Posts", color=C["txt3"], fontsize=8.5, labelpad=5)
        cls._setup_ax(ax, "⏰  Most Active Hours", "#00f5ff")
        return cls._save(fig)

    @classmethod
    def days_chart(cls, days: list) -> Image.Image:
        fig, ax = cls._base_fig(10.0, 3.4)
        mx = max(days) if any(days) else 1
        pi = days.index(mx) if mx > 0 else -1

        cls._gradient_bars(ax, range(7), days,
                           "#280040", "#c040f5", width=0.60, peak_idx=pi)

        avg = sum(days) / 7
        ax.axhline(avg, color="#4d8dff", linewidth=0.9, linestyle=":",
                   alpha=0.7, zorder=1)

        for i, v in enumerate(days):
            if v > 0:
                ax.text(i, v + mx*0.04, str(v), ha="center", va="bottom",
                        fontsize=9.5,
                        color="#c040f5" if i==pi else C["txt2"],
                        fontweight="bold" if i==pi else "normal")

        if pi >= 0:
            ax.annotate(f"Most active: {DAY_NAMES[pi]}",
                        xy=(pi, mx),
                        xytext=(pi+1.2 if pi<5 else pi-1.2, mx*0.82),
                        fontsize=8, color="#c040f5", fontweight="bold",
                        arrowprops=dict(arrowstyle="-|>", color="#c040f5",
                                        lw=0.8, mutation_scale=10))

        ax.set_xticks(range(7))
        ax.set_xticklabels(DAY_NAMES, color=C["txt2"], fontsize=10)
        ax.set_xlim(-0.5, 6.5)
        ax.set_ylim(0, mx * 1.25)
        ax.set_ylabel("Posts", color=C["txt3"], fontsize=8.5, labelpad=5)
        cls._setup_ax(ax, "📅  Most Active Days", "#c040f5")
        return cls._save(fig)

    @classmethod
    def hashtags_chart(cls, counter: Counter, top=15) -> Image.Image:
        items = counter.most_common(top)
        if not items: return None
        tags, counts = zip(*items)
        nh  = max(2.8, len(tags) * 0.26)
        fig, ax = cls._base_fig(10.0, nh)

        palette = [
            "#00f5ff","#4d8dff","#c040f5","#ff3d8a","#39ffac",
            "#ffd700","#ff9500","#00e5ff","#7c4dff","#64ffda",
            "#ff6e40","#40c4ff","#ea80fc","#ccff90","#80d8ff",
        ]
        for i, (tag, cnt) in enumerate(zip(tags[::-1], counts[::-1])):
            col  = palette[i % len(palette)]
            r,g,b = hex_to_rgb(col)
            pct  = cnt / counts[0]
            # Background track
            ax.barh(i, counts[0], height=0.62,
                    color=(r/255*0.07, g/255*0.07, b/255*0.07),
                    edgecolor="none")
            # Filled portion
            ax.barh(i, cnt, height=0.62,
                    color=(r/255, g/255, b/255, 0.85),
                    edgecolor=(r/255, g/255, b/255), linewidth=0.4)
            # Count label
            ax.text(cnt + counts[0]*0.012, i, f"{cnt}×",
                    va="center", fontsize=8, color=col,
                    fontweight="bold")

        ax.set_yticks(range(len(tags)))
        ax.set_yticklabels(list(tags)[::-1], fontsize=9, color=C["txt2"])
        ax.set_xlim(0, counts[0] * 1.18)
        ax.set_xlabel("Uses", color=C["txt3"], fontsize=8.5, labelpad=4)
        ax.xaxis.grid(True, color=C["ch_axis"], linewidth=0.3,
                      linestyle="--", alpha=0.8)
        ax.yaxis.grid(False)
        cls._setup_ax(ax, f"🏷  Top Hashtags  ({len(tags)} shown)", "#39ffac")
        fig.subplots_adjust(left=0.16, right=0.95, top=0.88, bottom=0.12)
        return cls._save(fig)

    @classmethod
    def engagement_chart(cls, medias) -> Image.Image:
        likes = [getattr(m,"like_count",0) or 0 for m in medias][::-1]
        cmts  = [getattr(m,"comment_count",0) or 0 for m in medias][::-1]
        if not likes: return None
        xs  = list(range(len(likes)))
        fig, ax = cls._base_fig(10.0, 3.4)

        # Gradient fill under likes
        r,g,b = hex_to_rgb("#00f5ff")
        grad = np.zeros((100, 1, 4))
        for row in range(100):
            grad[row,0] = [r/255,g/255,b/255, 0.06 + 0.08*(1-row/100)]
        ax.fill_between(xs, likes, alpha=0.0)   # reserve space
        if likes:
            mn, mx2 = min(likes), max(likes) or 1
            for ix in range(len(xs)):
                x  = xs[ix]; v = likes[ix]
                ax.bar(x, v, width=0.6, color=(r/255,g/255,b/255,0.10),
                       edgecolor="none", zorder=1)

        ax.plot(xs, likes, color="#00f5ff", linewidth=2.2,
                label="Likes", solid_capstyle="round",
                solid_joinstyle="round", zorder=4)
        ax.plot(xs, cmts,  color="#c040f5", linewidth=1.6,
                label="Comments", linestyle="--",
                solid_capstyle="round", zorder=4)

        # Peak like marker
        if likes:
            pi = likes.index(max(likes))
            ax.scatter([pi], [likes[pi]], s=55, color="#00f5ff",
                       zorder=5, edgecolors="white", linewidths=0.8)
            ax.annotate(f"Peak: {fmt_num(likes[pi])}",
                        xy=(pi, likes[pi]),
                        xytext=(pi + max(len(xs)*0.06, 2),
                                likes[pi] * 0.92),
                        fontsize=7.5, color="#00f5ff",
                        arrowprops=dict(arrowstyle="-",
                                        color="#00f5ff", lw=0.7))

        ax.set_xlabel("Post  (oldest → newest)",
                      color=C["txt3"], fontsize=8.5, labelpad=4)
        ax.set_ylabel("Count", color=C["txt3"], fontsize=8.5, labelpad=5)
        ax.legend(facecolor=C["ch_bg2"], edgecolor=C["ch_axis"],
                  labelcolor=C["txt2"], fontsize=8.5,
                  framealpha=0.9, loc="upper left")
        ax.set_xlim(0, max(len(likes)-1, 1))
        cls._setup_ax(ax, "📈  Engagement Over Time", "#39ffac")
        return cls._save(fig)


# ─────────────────────────────────────────────────────────────────────────────
#  Export engine
# ─────────────────────────────────────────────────────────────────────────────

class Exporter:

    @staticmethod
    def to_json(data, username) -> Path:
        fn = Path(f"instagrep_{username}_{datetime.now():%Y%m%d_%H%M%S}.json")
        fn.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
        return fn

    @staticmethod
    def to_excel(data, username) -> Path | None:
        if not OPENPYXL_OK: return None
        fn  = Path(f"instagrep_{username}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        wb  = openpyxl.Workbook()
        hf  = Font(bold=True, color="FFFFFF", size=11)
        hfl = PatternFill("solid", fgColor="1a237e")
        ha  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cf  = Font(color="eef0ff", size=10)
        af  = PatternFill("solid", fgColor="10101e")
        t   = Side(style="thin", color="22224a")
        bdr = Border(left=t, right=t, top=t, bottom=t)

        def hrow(ws, vals, r=1):
            for c,v in enumerate(vals,1):
                cell = ws.cell(row=r,column=c,value=v)
                cell.font=hf; cell.fill=hfl; cell.alignment=ha; cell.border=bdr

        def drow(ws, vals, r, alt=False):
            for c,v in enumerate(vals,1):
                cell = ws.cell(row=r,column=c,value=v)
                cell.font=cf
                cell.fill= PatternFill("solid", fgColor="0b0b18") if alt else af
                cell.alignment=Alignment(vertical="top",wrap_text=True)
                cell.border=bdr

        ws = wb.active; ws.title="Profile"; ws.sheet_view.showGridLines=False
        hrow(ws,["Field","Value"])
        prof = data.get("profile",{})
        for i,(f,v) in enumerate([
            ("Username",f"@{prof.get('username','')}"),("Full Name",prof.get("full_name","")),
            ("User ID",str(prof.get("pk",""))),("Followers",fmt_num(prof.get("follower_count",0))),
            ("Following",fmt_num(prof.get("following_count",0))),("Posts",fmt_num(prof.get("media_count",0))),
            ("Verified","Yes" if prof.get("is_verified") else "No"),
            ("Private","Yes" if prof.get("is_private") else "No"),
            ("Business","Yes" if prof.get("is_business") else "No"),
            ("Biography",prof.get("biography","")),
            ("External URL",str(prof.get("external_url",""))),
            ("Analyzed At",data.get("analyzed_at","")),
        ],2): drow(ws,[f,v],i,alt=i%2==0)
        ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=60

        ws2=wb.create_sheet("Posts"); ws2.sheet_view.showGridLines=False
        hrow(ws2,["#","Date","Type","Likes","Comments","Location","Caption"])
        for c,w in zip("ABCDEFG",[5,20,8,8,10,25,60]):
            ws2.column_dimensions[c].width=w
        tm={1:"Photo",2:"Video",8:"Album"}
        for i,p in enumerate(data.get("posts",[]),2):
            loc=p.get("location") or {}
            ls=", ".join(filter(None,[loc.get("name"),loc.get("city"),loc.get("country")]))
            drow(ws2,[i-1,p.get("taken_at","")[:19].replace("T"," "),
                      tm.get(p.get("type",1),"?"),p.get("likes",0),
                      p.get("comments",0),ls,p.get("caption","")[:200]],i,alt=i%2==0)

        ws3=wb.create_sheet("Activity"); ws3.sheet_view.showGridLines=False
        hrow(ws3,["Metric","Value"])
        act=data.get("activity",{})
        for i,(f,v) in enumerate([
            ("Avg Posts/Day",act.get("avg_posts_per_day","")),
            ("Peak Hour",act.get("peak_hour","")),("Peak Day",act.get("peak_day","")),
            ("Avg Likes",act.get("avg_likes","")),("Avg Comments",act.get("avg_comments","")),
            ("Top Hashtags",", ".join(act.get("top_hashtags",[])[:10])),
        ],2): drow(ws3,[f,str(v)],i,alt=i%2==0)
        ws3.column_dimensions["A"].width=22; ws3.column_dimensions["B"].width=50
        wb.save(fn); return fn

    @staticmethod
    def to_html(data, username) -> Path:
        fn   = Path(f"instagrep_{username}_{datetime.now():%Y%m%d_%H%M%S}.html")
        prof = data.get("profile",{}); posts=data.get("posts",[]); act=data.get("activity",{})
        def esc(s): return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        rows = "".join(f"""<tr><td>{esc(p.get("taken_at","")[:10])}</td>
          <td>{esc({1:"📷",2:"🎬",8:"📚"}.get(p.get("type",1),"?"))}</td>
          <td>{fmt_num(p.get("likes",0))}</td><td>{fmt_num(p.get("comments",0))}</td>
          <td>{esc(", ".join(filter(None, [((p.get("location") or {}).get(k, "")) for k in ("name", "city", "country")])))}</td>
          <td class="cap">{esc((p.get("caption","") or "")[:180])}</td></tr>"""
          for p in posts[:50])
        fn.write_text(f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><title>InstaGrep — @{esc(username)}</title>
<style>:root{{--bg:#07070f;--bg2:#0b0b18;--card:#14142a;--border:#22224a;
  --txt:#eef0ff;--txt2:#9ba8d8;--txt3:#565e9e;--cyan:#00f5ff;--green:#4dffb0;--purple:#c855f5}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--txt);font-family:'Courier New',monospace;font-size:13px}}
.hdr{{background:var(--bg2);border-bottom:1px solid var(--border);padding:20px 32px;
      display:flex;align-items:center;justify-content:space-between}}
.hdr h1{{font-size:20px;color:var(--cyan)}} .sub{{color:var(--txt3);font-size:11px}}
.body{{padding:20px 32px;max-width:1400px;margin:0 auto}}
.stats{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:20px}}
.stat{{background:var(--card);border:1px solid var(--border);border-radius:10px;
       padding:14px;text-align:center}}
.sv{{font-size:24px;font-weight:bold;color:var(--cyan)}}
.sl{{font-size:10px;color:var(--txt3);margin-top:4px}}
.card{{background:var(--card);border:1px solid var(--border);border-radius:10px;
       padding:16px;margin-bottom:16px}}
.ct{{font-size:12px;font-weight:bold;color:var(--cyan);border-bottom:1px solid var(--border);
     padding-bottom:8px;margin-bottom:12px}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
th{{background:#1a237e;color:#fff;padding:8px 10px;text-align:left}}
td{{padding:7px 10px;border-bottom:1px solid var(--border);vertical-align:top}}
tr:nth-child(even) td{{background:var(--bg2)}}
.cap{{max-width:380px;color:var(--txt2)}}
.footer{{text-align:center;padding:16px;color:var(--txt3);font-size:10px;
         border-top:1px solid var(--border)}}
</style></head><body>
<div class="hdr"><div><h1>⚡ InstaGrep v4.0</h1>
<div class="sub">Instagram OSINT Analysis — @{esc(username)}</div></div>
<div class="sub">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} • {len(posts)} posts</div></div>
<div class="body">
<div class="stats">
{''.join(f'<div class="stat"><div class="sv">{v}</div><div class="sl">{l}</div></div>'
  for v,l in [(fmt_num(prof.get("follower_count",0)),"Followers"),
              (fmt_num(prof.get("following_count",0)),"Following"),
              (fmt_num(prof.get("media_count",0)),"Posts"),
              (fmt_num(act.get("avg_likes",0)),"Avg Likes"),
              (fmt_num(act.get("avg_comments",0)),"Avg Comments"),
              (str(act.get("peak_hour","—")),"Peak Hour")])}
</div>
<div class="card"><div class="ct">Posts</div>
<table><thead><tr><th>Date</th><th>Type</th><th>Likes</th><th>Comments</th><th>Location</th><th>Caption</th></tr></thead>
<tbody>{rows}</tbody></table></div>
</div><div class="footer">InstaGrep v4.0 — {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
</body></html>""", encoding="utf-8")
        return fn

    @staticmethod
    def to_pdf(data, username, chart_paths=None) -> Path | None:
        if not REPORTLAB_OK: return None
        fn   = Path(f"instagrep_{username}_{datetime.now():%Y%m%d_%H%M%S}.pdf")
        prof = data.get("profile",{}); posts=data.get("posts",[]); act=data.get("activity",{})
        doc  = SimpleDocTemplate(str(fn),pagesize=A4,
                                 leftMargin=0.6*inch,rightMargin=0.6*inch,
                                 topMargin=0.6*inch,bottomMargin=0.6*inch)
        CARD=colors.HexColor("#14142a"); CYAN=colors.HexColor("#00f5ff")
        TXT=colors.HexColor("#eef0ff"); TXT2=colors.HexColor("#9ba8d8")
        BL=colors.HexColor("#1a237e"); BDR=colors.HexColor("#22224a")
        ts = ParagraphStyle; story=[]
        T = ts("t",fontSize=20,textColor=CYAN,fontName="Helvetica-Bold",spaceAfter=4,alignment=TA_CENTER)
        S = ts("s",fontSize=10,textColor=TXT2,fontName="Helvetica",spaceAfter=16,alignment=TA_CENTER)
        H = ts("h",fontSize=13,textColor=CYAN,fontName="Helvetica-Bold",spaceAfter=6)
        SM= ts("sm",fontSize=8,textColor=TXT2,fontName="Helvetica")
        story += [Paragraph(f"InstaGrep v4.0  —  @{username}",T),
                  Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  •  {len(posts)} posts",S),
                  HRFlowable(width="100%",thickness=1,color=BDR,spaceAfter=12)]
        it = TableStyle([("BACKGROUND",(0,0),(-1,-1),CARD),("BACKGROUND",(0,0),(0,-1),BL),
             ("TEXTCOLOR",(0,0),(0,-1),CYAN),("TEXTCOLOR",(1,0),(1,-1),TXT),
             ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(1,0),(1,-1),"Helvetica"),
             ("FONTSIZE",(0,0),(-1,-1),9),("GRID",(0,0),(-1,-1),0.5,BDR),
             ("ROWBACKGROUNDS",(1,0),(1,-1),[CARD,colors.HexColor("#10101e")]),
             ("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,0),(-1,-1),6),
             ("BOTTOMPADDING",(0,0),(-1,-1),6),("LEFTPADDING",(0,0),(-1,-1),8)])
        story.append(Paragraph("Account Information",H))
        story.append(Table([["Full Name",prof.get("full_name","—")],
             ["Username",f"@{prof.get('username','—')}"],["User ID",str(prof.get("pk","—"))],
             ["Biography",(prof.get("biography") or "—")[:200]]],
             colWidths=[1.3*inch,5.5*inch],style=it))
        story.append(Spacer(1,12))
        if chart_paths:
            for ttl,path in chart_paths.items():
                if path and Path(path).exists():
                    story += [Paragraph(ttl,H),
                              RLImage(str(path),width=6.4*inch,height=2.3*inch),
                              Spacer(1,6)]
        story.append(Paragraph(f"Posts ({len(posts)})",H))
        ph=["Date","Type","Likes","Comments","Location","Caption"]
        pr=[ph]+[[p.get("taken_at","")[:10],
             {1:"Photo",2:"Video",8:"Album"}.get(p.get("type",1),"?"),
             fmt_num(p.get("likes",0)),fmt_num(p.get("comments",0)),
             (", ".join(filter(None,[(p.get("location") or {}).get(k,"") for k in ["name","city","country"]]))or"—")[:25],
             (p.get("caption","") or "")[:75]] for p in posts[:50]]
        pts=TableStyle([("BACKGROUND",(0,0),(-1,0),BL),("TEXTCOLOR",(0,0),(-1,0),colors.white),
             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.5),
             ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("TEXTCOLOR",(0,1),(-1,-1),TXT),
             ("ROWBACKGROUNDS",(0,1),(-1,-1),[CARD,colors.HexColor("#10101e")]),
             ("GRID",(0,0),(-1,-1),0.4,BDR),("TOPPADDING",(0,0),(-1,-1),4),
             ("BOTTOMPADDING",(0,0),(-1,-1),4),("LEFTPADDING",(0,0),(-1,-1),5),
             ("VALIGN",(0,0),(-1,-1),"TOP")])
        story += [Table(pr,colWidths=[0.75*inch,0.5*inch,0.5*inch,0.6*inch,1.0*inch,3.2*inch],
                        style=pts,repeatRows=1),
                  Spacer(1,8),HRFlowable(width="100%",thickness=1,color=BDR),
                  Paragraph(f"InstaGrep v4.0 — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",SM)]
        doc.build(story)
        return fn


# ─────────────────────────────────────────────────────────────────────────────
#  Toast notification widget
# ─────────────────────────────────────────────────────────────────────────────

class Toast:
    """Slide-in, auto-dismiss toast overlay in top-right corner."""

    _active: list = []

    @classmethod
    def show(cls, root, message: str, kind="info", duration=3200):
        col = {"ok": C["s_ok"], "warn": C["s_warn"],
               "err": C["s_err"], "info": C["cyan"]}.get(kind, C["cyan"])
        icon= {"ok":"✅","warn":"⚠","err":"❌","info":"ℹ"}.get(kind,"ℹ")

        win = tk.Toplevel(root)
        win.overrideredirect(True)
        win.configure(bg=C["card2"])
        win.attributes("-alpha", 0.0)
        win.attributes("-topmost", True)

        fr = tk.Frame(win, bg=C["card2"], bd=0)
        fr.pack(padx=1, pady=1)
        tk.Label(fr, text=f"  {icon}  {message}  ",
                 bg=C["card2"], fg=col,
                 font=("Courier New", 11),
                 padx=12, pady=9).pack()
        # Accent line at top
        tk.Frame(fr, bg=col, height=2).pack(fill="x", side="top")

        win.update_idletasks()
        rw = root.winfo_width()
        rx = root.winfo_x()
        ry = root.winfo_y()
        tw = win.winfo_reqwidth()
        th = win.winfo_reqheight()

        y_start = ry + 80
        offset  = len([w for w in cls._active if w.winfo_exists()]) * (th + 8)
        x_pos   = rx + rw - tw - 16
        y_pos   = y_start + offset

        win.geometry(f"+{x_pos}+{y_pos}")
        cls._active.append(win)

        # Fade in
        def fade_in(a=0.0):
            if not win.winfo_exists(): return
            a = min(a + 0.08, 0.95)
            win.attributes("-alpha", a)
            if a < 0.95: root.after(16, fade_in, a)

        # Fade out + destroy
        def fade_out(a=0.95):
            if not win.winfo_exists(): return
            a = max(a - 0.06, 0.0)
            win.attributes("-alpha", a)
            if a > 0:
                root.after(16, fade_out, a)
            else:
                try: cls._active.remove(win)
                except: pass
                win.destroy()

        root.after(0, fade_in)
        root.after(duration, fade_out)


# ─────────────────────────────────────────────────────────────────────────────
#  Main application
# ─────────────────────────────────────────────────────────────────────────────

BASE = TkinterDnD.Tk if DND_OK else ctk.CTk

class App(BASE):

    def __init__(self):
        super().__init__()
        if not DND_OK:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")

        self.title(APP_TITLE)
        self.geometry("1460x900")
        self.minsize(1100, 720)
        self.configure(fg_color=C["bg"])

        self._q             = queue.Queue()
        self._data          = {}
        self._cl            = None
        self._sessions      = SessionManager()
        self._recent        = RecentSearches()
        self._dl_mgr        = None
        self._chart_paths   = {}
        self._medias_cache  = []
        self._analyzing     = False
        self._tw_job        = None
        self._tw_target     = ""
        self._tw_current    = ""
        self._count_jobs    = {}   # widget → after-id
        self._recent_popup  = None
        self._pulse_state   = True
        self._cached_follower_count = 0

        self._build_ui()
        self._bind_keys()
        self._poll()
        self._pulse_dot()

        if self._sessions.has():
            self.after(300, self._activate_saved)
        else:
            self.after(400, self._probe_async)

        if DND_OK:
            self.drop_target_register(DND_FILES)
            self.dnd_bind("<<Drop>>", self._on_drop)

    # ══════════════════════════════════════════════════════════════════════════
    #  Keyboard bindings
    # ══════════════════════════════════════════════════════════════════════════

    def _bind_keys(self):
        self.bind("<Return>",        lambda _: self._start())
        self.bind("<Control-Return>",lambda _: self._start())
        self.bind("<Control-s>",     lambda _: self._save_json())
        self.bind("<Control-S>",     lambda _: self._save_json())
        self.bind("<Control-d>",     lambda _: self._dl_start())
        self.bind("<Control-D>",     lambda _: self._dl_start())
        self.bind("<Control-e>",     lambda _: self._export_pdf())
        self.bind("<Control-E>",     lambda _: self._export_pdf())
        self.bind("<Control-h>",     lambda _: self._toggle_recent())
        self.bind("<Control-H>",     lambda _: self._toggle_recent())
        self.bind("?",               lambda _: self._show_shortcuts())
        self._entry.bind("<Return>", lambda _: self._start())
        self._entry.bind("<Down>",   lambda _: self._toggle_recent())
        self.bind("<Escape>",        lambda _: self._close_recent())

    def _show_shortcuts(self):
        w = ctk.CTkToplevel(self)
        w.title("Keyboard Shortcuts"); w.geometry("440x360")
        w.configure(fg_color=C["card"]); w.grab_set()
        ctk.CTkLabel(w, text="⌨  Keyboard Shortcuts",
                     font=ctk.CTkFont(size=16,weight="bold"),
                     text_color=C["cyan"]).pack(pady=(20,10))
        ctk.CTkFrame(w,height=1,fg_color=C["border"]).pack(fill="x",padx=20)
        for key,desc in [("Enter","Analyze username"),("Ctrl+S","Save JSON"),
                         ("Ctrl+D","Download media"),("Ctrl+E","Export PDF"),
                         ("Ctrl+H","Toggle recent searches"),
                         ("↓ (in box)","Open recent searches"),
                         ("Esc","Close panels"),("?","This cheatsheet")]:
            r=ctk.CTkFrame(w,fg_color="transparent"); r.pack(fill="x",padx=24,pady=5)
            ctk.CTkLabel(r,text=key,
                         font=ctk.CTkFont(family="Courier New",size=11,weight="bold"),
                         text_color=C["yellow"],fg_color=C["bg3"],
                         corner_radius=4,width=140,anchor="center").pack(side="left")
            ctk.CTkLabel(r,text=desc,font=ctk.CTkFont(size=11),
                         text_color=C["txt2"],anchor="w").pack(side="left",padx=12)
        ctk.CTkButton(w,text="Close",command=w.destroy,
                      fg_color=C["blue"],hover_color=C["blue2"],width=100).pack(pady=16)

    # ══════════════════════════════════════════════════════════════════════════
    #  Pulse animation (status dot)
    # ══════════════════════════════════════════════════════════════════════════

    def _pulse_dot(self):
        if not self._analyzing:
            self._dot.configure(text_color=C["s_ok"])
        else:
            col = C["s_warn"] if self._pulse_state else C["bg2"]
            self._dot.configure(text_color=col)
            self._pulse_state = not self._pulse_state
        self.after(600, self._pulse_dot)

    # ══════════════════════════════════════════════════════════════════════════
    #  DnD
    # ══════════════════════════════════════════════════════════════════════════

    def _on_drop(self, event):
        path = event.data.strip().strip("{}")
        if path.lower().endswith(".json"):
            threading.Thread(target=self._load_cookie_file,
                             args=(path,), daemon=True).start()
        else:
            Toast.show(self, "Drop a cookies.json file", "warn")

    # ══════════════════════════════════════════════════════════════════════════
    #  UI construction
    # ══════════════════════════════════════════════════════════════════════════

    def _build_ui(self):
        self._build_header()
        self._build_session_bar()
        self._build_tabs()
        self._build_status_bar()

    # ── Header ────────────────────────────────────────────────────────────────

    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=C["bg2"], height=74, corner_radius=0)
        hdr.pack(fill="x"); hdr.pack_propagate(False)

        # Logo
        left = ctk.CTkFrame(hdr, fg_color="transparent")
        left.pack(side="left", padx=20)
        ctk.CTkLabel(left, text="⚡",
                     font=ctk.CTkFont(size=26),
                     text_color=C["cyan"]).pack(side="left", padx=(0,6))
        name_f = ctk.CTkFrame(left, fg_color="transparent")
        name_f.pack(side="left")
        ctk.CTkLabel(name_f, text="INSTAGREP",
                     font=ctk.CTkFont(family="Courier New", size=20, weight="bold"),
                     text_color=C["cyan"]).pack(anchor="w")
        ctk.CTkLabel(name_f, text="v4.0  —  Instagram OSINT",
                     font=ctk.CTkFont(size=10), text_color=C["txt4"]).pack(anchor="w")

        # Right: search area
        right = ctk.CTkFrame(hdr, fg_color="transparent")
        right.pack(side="right", padx=20)

        # Recent searches button
        self._recent_btn = ctk.CTkButton(
            right, text="▼", command=self._toggle_recent,
            width=36, height=44,
            font=ctk.CTkFont(size=13),
            fg_color=C["card2"], hover_color=C["card3"],
            text_color=C["txt3"], corner_radius=8)
        self._recent_btn.pack(side="left", padx=(0,4))

        self._uv    = ctk.StringVar()
        self._entry = ctk.CTkEntry(
            right, textvariable=self._uv,
            placeholder_text="Username or profile URL …",
            width=280, height=44,
            font=ctk.CTkFont(size=13),
            fg_color=C["card"], border_color=C["border"],
            text_color=C["txt"], corner_radius=8)
        self._entry.pack(side="left", padx=(0,8))

        self._btn = ctk.CTkButton(
            right, text="⚡  ANALYZE", command=self._start,
            width=155, height=44,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=C["blue"], hover_color=C["blue2"],
            corner_radius=8)
        self._btn.pack(side="left")

        # Action icon buttons
        for icon, tip, cmd in (
            ("⬇","Download (Ctrl+D)",   self._dl_start),
            ("📄","PDF (Ctrl+E)",        self._export_pdf),
            ("📊","Excel",               self._export_excel),
            ("🌐","HTML",                self._export_html),
        ):
            b = ctk.CTkButton(right, text=icon, command=cmd,
                              width=42, height=44,
                              font=ctk.CTkFont(size=16),
                              fg_color=C["card"], hover_color=C["card3"],
                              text_color=C["txt3"], corner_radius=8)
            b.pack(side="left", padx=3)
            self._add_hover_glow(b)

    def _add_hover_glow(self, btn):
        def on_enter(e):
            btn.configure(border_color=C["cyan"], border_width=1)
        def on_leave(e):
            btn.configure(border_width=0)
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

    # ── Recent searches popup ─────────────────────────────────────────────────

    def _toggle_recent(self):
        if self._recent_popup and self._recent_popup.winfo_exists():
            self._close_recent()
        else:
            self._open_recent()

    def _close_recent(self):
        if self._recent_popup and self._recent_popup.winfo_exists():
            self._recent_popup.destroy()
        self._recent_popup = None

    def _open_recent(self):
        items = self._recent.get_all()
        self.update_idletasks()

        # Position below the entry
        ex = self._entry.winfo_rootx()
        ey = self._entry.winfo_rooty() + self._entry.winfo_height() + 6
        w  = 400
        row_h    = 58
        header_h = 46
        target_h = header_h + max(min(len(items), 8) * row_h, 56) + 2

        pop = tk.Toplevel(self)
        pop.overrideredirect(True)
        pop.configure(bg=C["border"])          # thin border effect
        pop.attributes("-topmost", True)
        pop.geometry(f"{w}x0+{ex}+{ey}")       # start height 0 for animation
        self._recent_popup = pop

        # ── Outer container ───────────────────────────────────────────────────
        outer = tk.Frame(pop, bg=C["card2"])
        outer.pack(fill="both", expand=True, padx=1, pady=1)

        # ── Header ────────────────────────────────────────────────────────────
        hf = tk.Frame(outer, bg=C["bg2"], height=header_h)
        hf.pack(fill="x"); hf.pack_propagate(False)

        # Clock icon + title
        tk.Label(hf, text="🕐", bg=C["bg2"], fg=C["cyan"],
                 font=("", 13)).pack(side="left", padx=(12, 4), pady=0)
        tk.Label(hf, text="Recent Searches",
                 bg=C["bg2"], fg=C["txt2"],
                 font=("Courier New", 11, "bold")).pack(side="left")

        # Count badge
        if items:
            badge_bg = C["blue3"]
            tk.Label(hf, text=f" {len(items)} ",
                     bg=badge_bg, fg=C["txt"],
                     font=("Courier New", 9, "bold"),
                     padx=4).pack(side="left", padx=8)

        if items:
            btn = tk.Button(hf, text="✕ Clear",
                            bg=C["bg2"], fg=C["red"],
                            font=("Courier New", 9), bd=0,
                            cursor="hand2", activebackground=C["red2"],
                            activeforeground="#fff",
                            command=self._clear_recent)
            btn.pack(side="right", padx=12)

        # Separator
        tk.Frame(outer, bg=C["border"], height=1).pack(fill="x")

        if not items:
            tk.Label(outer, text="  No recent searches yet",
                     bg=C["card2"], fg=C["txt3"],
                     font=("Courier New", 11), pady=20).pack()
        else:
            sc_frame = tk.Frame(outer, bg=C["card2"])
            sc_frame.pack(fill="both", expand=True)

            for idx, item in enumerate(items):
                username  = item.get("username", "")
                full_name = item.get("full_name", "")
                ts_str    = item.get("timestamp", "")
                try:
                    dt  = datetime.fromisoformat(ts_str)
                    ago = time_ago(dt)
                except: ago = "—"

                # Row frame
                row = tk.Frame(sc_frame, bg=C["card2"],
                               cursor="hand2", height=row_h)
                row.pack(fill="x"); row.pack_propagate(False)

                # Initials avatar
                av_img = make_initials_avatar(
                    username[:1] if username else "?",
                    size=34, hue_seed=idx)
                av_ci  = ctk.CTkImage(av_img, av_img, size=(34, 34))
                av_lbl = tk.Label(row, bg=C["card2"])
                av_lbl.configure(image=av_ci)   # type: ignore
                av_lbl._img = av_ci
                av_lbl.pack(side="left", padx=(12, 8), pady=12)

                # Text column
                txt_f = tk.Frame(row, bg=C["card2"])
                txt_f.pack(side="left", fill="y", expand=True, pady=8)
                tk.Label(txt_f, text=f"@{username}",
                         bg=C["card2"], fg=C["cyan"],
                         font=("Courier New", 11, "bold"),
                         anchor="w").pack(anchor="w")
                if full_name:
                    tk.Label(txt_f, text=full_name[:28],
                             bg=C["card2"], fg=C["txt3"],
                             font=("Courier New", 9),
                             anchor="w").pack(anchor="w")

                # Time-ago badge (right)
                ago_lbl = tk.Label(row, text=f"  {ago}  ",
                                   bg=C["card3"], fg=C["txt4"],
                                   font=("Courier New", 8),
                                   padx=4, pady=2)
                ago_lbl.pack(side="right", padx=12)

                # Separator line
                tk.Frame(sc_frame, bg=C["border"], height=1).pack(fill="x")

                # Click handler
                def _click(u=username):
                    self._uv.set(u)
                    self._close_recent()
                    self._start()

                # Hover highlight
                all_widgets = [row, txt_f, av_lbl]
                all_labels  = list(txt_f.winfo_children()) + [av_lbl]

                def _hin(e, r=row, tf=txt_f, al=all_labels):
                    r.configure(bg=C["card3"]); tf.configure(bg=C["card3"])
                    for ww in al:
                        try: ww.configure(bg=C["card3"])
                        except: pass

                def _hout(e, r=row, tf=txt_f, al=all_labels):
                    r.configure(bg=C["card2"]); tf.configure(bg=C["card2"])
                    for ww in al:
                        try: ww.configure(bg=C["card2"])
                        except: pass

                for wdg in [row, txt_f] + list(txt_f.winfo_children()):
                    wdg.bind("<Button-1>", lambda e, c=_click: c())
                    wdg.bind("<Enter>", _hin)
                    wdg.bind("<Leave>", _hout)
                av_lbl.bind("<Button-1>", lambda e, c=_click: c())

        pop.focus_force()
        pop.bind("<FocusOut>",
                 lambda e: self.after(120, self._maybe_close_recent))

        # ── Slide-down animation ──────────────────────────────────────────────
        def _slide(cur=0):
            if not pop.winfo_exists(): return
            step = int(target_h * 0.18)
            cur  = min(cur + step, target_h)
            pop.geometry(f"{w}x{cur}+{ex}+{ey}")
            if cur < target_h:
                pop.after(14, _slide, cur)

        pop.after(10, _slide)

    def _maybe_close_recent(self):
        if self._recent_popup and self._recent_popup.winfo_exists():
            try:
                focused = self.focus_get()
                if focused and str(focused).startswith(str(self._recent_popup)):
                    return
            except: pass
            self._close_recent()

    def _clear_recent(self):
        self._recent.clear()
        self._close_recent()
        Toast.show(self, "Recent searches cleared", "info")

    # ── Session bar ───────────────────────────────────────────────────────────

    def _build_session_bar(self):
        bar = ctk.CTkFrame(self, fg_color=C["bg3"], height=42, corner_radius=0)
        bar.pack(fill="x"); bar.pack_propagate(False)
        ctk.CTkFrame(bar,height=1,fg_color=C["border"]).place(relx=0,rely=0,relwidth=1)

        ctk.CTkLabel(bar, text="🔐  SESSION",
                     font=ctk.CTkFont(size=10,weight="bold"),
                     text_color=C["txt4"]).pack(side="left",padx=(14,10))
        ctk.CTkFrame(bar,width=1,fg_color=C["border"]).pack(side="left",fill="y",pady=8)

        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.pack(side="left", padx=12)
        self._ck_dot = ctk.CTkLabel(inner, text="●",
                                    font=ctk.CTkFont(size=14),
                                    text_color=C["s_warn"])
        self._ck_dot.pack(side="left", padx=(0,6))
        self._ck_sv = ctk.StringVar(value="No session — import cookies.json")
        ctk.CTkLabel(inner, textvariable=self._ck_sv,
                     font=ctk.CTkFont(size=11),
                     text_color=C["txt2"]).pack(side="left")

        for txt,cmd in (("📂  Import cookies.json",self._import_cookie),
                        ("🔄  Re-scan",self._probe_async)):
            ctk.CTkButton(bar,text=txt,command=cmd,
                          width=145,height=26,font=ctk.CTkFont(size=10),
                          fg_color=C["card"],hover_color=C["card3"],
                          text_color=C["txt2"]).pack(side="right",padx=4)

        self._sess_var = ctk.StringVar(value="Sessions …")
        self._sess_menu = ctk.CTkOptionMenu(
            bar, variable=self._sess_var,
            values=self._sessions.list() or ["—"],
            command=self._switch_session,
            width=150,height=26,font=ctk.CTkFont(size=10),
            fg_color=C["card"],button_color=C["card3"],text_color=C["txt2"])
        self._sess_menu.pack(side="right",padx=4)
        ctk.CTkLabel(bar,text="Sessions:",
                     font=ctk.CTkFont(size=10),
                     text_color=C["txt3"]).pack(side="right",padx=(8,0))

    # ── Status bar ────────────────────────────────────────────────────────────

    def _build_status_bar(self):
        bar = ctk.CTkFrame(self, fg_color=C["bg2"], height=28, corner_radius=0)
        bar.pack(fill="x", side="bottom"); bar.pack_propagate(False)

        self._dot = ctk.CTkLabel(bar, text="●",
                                 text_color=C["s_ok"],
                                 font=ctk.CTkFont(size=13))
        self._dot.pack(side="left", padx=(10,4))

        self._sv = ctk.StringVar(value="Ready — import cookies.json, then enter a target username")
        ctk.CTkLabel(bar, textvariable=self._sv,
                     font=ctk.CTkFont(size=11),
                     text_color=C["txt2"]).pack(side="left")

        self._prog = ctk.CTkProgressBar(bar, width=190, mode="indeterminate",
                                        fg_color=C["bg3"],
                                        progress_color=C["blue"])

        ctk.CTkLabel(bar, text="Press  ?  for shortcuts",
                     font=ctk.CTkFont(size=10),
                     text_color=C["txt4"]).pack(side="right", padx=12)

    # ══════════════════════════════════════════════════════════════════════════
    #  Tabs
    # ══════════════════════════════════════════════════════════════════════════

    def _build_tabs(self):
        self._tabs = ctk.CTkTabview(
            self,
            fg_color=C["bg2"],
            segmented_button_fg_color=C["bg3"],
            segmented_button_selected_color=C["blue2"],
            segmented_button_selected_hover_color=C["blue"],
            segmented_button_unselected_color=C["bg3"],
            segmented_button_unselected_hover_color=C["card2"],
            text_color=C["txt"],
        )
        self._tabs.pack(fill="both", expand=True, padx=8, pady=(4,0))
        for name in ("🪪  Profile", "📬  Contact", "📍  Locations",
                     "🗂  Posts", "📊  Activity", "⬇  Downloads",
                     "📤  Export", "{}  Raw"):
            self._tabs.add(name)

        self._build_profile_tab()
        self._build_contact_tab()
        self._build_locations_tab()
        self._build_posts_tab()
        self._build_activity_tab()
        self._build_downloads_tab()
        self._build_export_tab()
        self._build_raw_tab()

    # ══════════════════════════════════════════════════════════════════════════
    #  PROFILE TAB  — v4 glassmorphism redesign
    # ══════════════════════════════════════════════════════════════════════════

    def _build_profile_tab(self):
        t  = self._tabs.tab("🪪  Profile")
        sc = ctk.CTkScrollableFrame(t, fg_color="transparent",
                                    scrollbar_button_color=C["border"],
                                    scrollbar_button_hover_color=C["border2"])
        sc.pack(fill="both", expand=True, padx=2, pady=2)

        # ── HERO BANNER  (canvas gradient) ───────────────────────────────────
        self._hero_canvas = tk.Canvas(sc, height=190, highlightthickness=0,
                                      bg=C["bg"])
        self._hero_canvas.pack(fill="x", padx=4, pady=(4, 0))
        self._hero_canvas.bind("<Configure>", self._draw_hero_bg)
        self._draw_hero_bg()

        # Avatar on top-left of hero (positioned absolutely)
        self._pic_frame = ctk.CTkFrame(sc, fg_color=C["card"],
                                        corner_radius=20, width=196, height=196)
        self._pic_frame.pack_propagate(False)
        self._pic_frame.place(in_=self._hero_canvas, x=18, y=10)
        self._pic_lbl = ctk.CTkLabel(self._pic_frame, text="")
        self._pic_lbl.pack(expand=True)
        self._render_placeholder_avatar()

        # Name / username over hero (canvas text)
        self._hero_name_var = ctk.StringVar(value="")
        self._hero_user_var = ctk.StringVar(value="")

        # ── PROFILE IDENTITY BAR  (below hero) ───────────────────────────────
        id_bar = ctk.CTkFrame(sc, fg_color=C["card"], corner_radius=0,
                               border_color=C["border"], border_width=1)
        id_bar.pack(fill="x", padx=4)

        id_inner = ctk.CTkFrame(id_bar, fg_color="transparent")
        id_inner.pack(fill="x", padx=220, pady=(10,12))  # left pad clears avatar

        # Name row
        name_row = ctk.CTkFrame(id_inner, fg_color="transparent")
        name_row.pack(anchor="w")
        self._lbl_fullname = ctk.CTkLabel(
            name_row, text="—",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=C["txt"])
        self._lbl_fullname.pack(side="left")

        self._lbl_uid = ctk.CTkLabel(
            name_row, text="",
            font=ctk.CTkFont(size=10),
            text_color=C["txt4"])
        self._lbl_uid.pack(side="left", padx=(12,0))

        # Username + badge row
        un_row = ctk.CTkFrame(id_inner, fg_color="transparent")
        un_row.pack(anchor="w", pady=(2,8))
        self._lbl_username = ctk.CTkLabel(
            un_row, text="@—",
            font=ctk.CTkFont(family="Courier New", size=14, weight="bold"),
            text_color=C["cyan"])
        self._lbl_username.pack(side="left")

        self._badge_frame = ctk.CTkFrame(un_row, fg_color="transparent")
        self._badge_frame.pack(side="left", padx=(14, 0))

        # Info grid  (4 small pill widgets in 2×2)
        ig = ctk.CTkFrame(id_inner, fg_color="transparent")
        ig.pack(anchor="w")
        ig.columnconfigure([0,1,2,3], weight=0)

        self._vtype  = ctk.StringVar(value="—")
        self._vverif = ctk.StringVar(value="—")
        self._vurl   = ctk.StringVar(value="—")
        self._vcat   = ctk.StringVar(value="—")

        for col, (lbl, var, acc) in enumerate((
            ("Type",     self._vtype,  C["blue"]),
            ("Verified", self._vverif, C["green"]),
            ("URL",      self._vurl,   C["cyan"]),
            ("Category", self._vcat,   C["txt2"]),
        )):
            pill = ctk.CTkFrame(ig, fg_color=C["card2"],
                                corner_radius=8, border_color=C["border"],
                                border_width=1)
            pill.grid(row=0, column=col, padx=(0,8), sticky="w")
            ctk.CTkLabel(pill, text=lbl,
                         font=ctk.CTkFont(size=9, weight="bold"),
                         text_color=C["txt4"]).pack(
                             side="left", padx=(10,4), pady=6)
            ctk.CTkLabel(pill, textvariable=var,
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color=acc).pack(side="left", padx=(0,10))
            self._add_hover_glow(pill)

        # ── STAT CARDS  (glassmorphism) ───────────────────────────────────────
        stats_outer = ctk.CTkFrame(sc, fg_color="transparent")
        stats_outer.pack(fill="x", padx=4, pady=(10,0))
        for i in range(3):
            stats_outer.columnconfigure(i, weight=1)

        stat_cfgs = [
            ("followers", "👥  Followers", C["cyan"],   C["cyan_dim"]),
            ("following", "🔗  Following", C["purple"], C["purple_dim"]),
            ("posts",     "📷  Posts",     C["green"],  C["green_dim"]),
        ]
        self._stat_labels    = {}
        self._stat_canvas    = {}
        self._stat_sparkline = {}

        for col, (key, title, color, bg_dim) in enumerate(stat_cfgs):
            # Outer glass frame
            card = ctk.CTkFrame(stats_outer, fg_color=bg_dim,
                                corner_radius=18,
                                border_color=color, border_width=1)
            card.grid(row=0, column=col,
                      padx=(0, 10) if col < 2 else (0, 0), sticky="ew")

            # Canvas for background gradient + sparkline
            cv = tk.Canvas(card, height=110, bg=bg_dim,
                           highlightthickness=0)
            cv.pack(fill="x")
            self._stat_canvas[key] = (cv, color, bg_dim)

            # Top accent bar
            ctk.CTkFrame(card, height=3, fg_color=color,
                         corner_radius=0).pack(fill="x")

            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=16, pady=(8, 14))

            ctk.CTkLabel(inner, text=title,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=color).pack(anchor="w")

            num_lbl = ctk.CTkLabel(inner, text="—",
                                   font=ctk.CTkFont(size=36, weight="bold"),
                                   text_color=color)
            num_lbl.pack(anchor="w", pady=(4, 0))

            # Sparkline placeholder label
            sp_lbl = ctk.CTkLabel(inner, text="",
                                  font=ctk.CTkFont(size=1))
            sp_lbl.pack(anchor="w")

            self._stat_labels[key]    = (num_lbl, color)
            self._stat_sparkline[key] = sp_lbl

            # Hover lift effect
            def _enter(e, c=card, cl=color):
                c.configure(border_width=2, border_color=cl)
            def _leave(e, c=card, col_=color, bg_=bg_dim):
                c.configure(border_width=1, border_color=col_)
            card.bind("<Enter>", _enter)
            card.bind("<Leave>", _leave)

        # ── ENGAGEMENT RATE RIBBON ────────────────────────────────────────────
        er_frame = ctk.CTkFrame(sc, fg_color=C["card"],
                                corner_radius=14,
                                border_color=C["border"], border_width=1)
        er_frame.pack(fill="x", padx=4, pady=(10, 0))

        er_inner = ctk.CTkFrame(er_frame, fg_color="transparent")
        er_inner.pack(fill="x", padx=18, pady=12)

        ctk.CTkLabel(er_inner, text="⚡",
                     font=ctk.CTkFont(size=18),
                     text_color=C["yellow"]).pack(side="left", padx=(0, 6))
        ctk.CTkLabel(er_inner, text="Engagement Rate",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C["txt3"]).pack(side="left")

        self._ver = ctk.StringVar(value="—")
        ctk.CTkLabel(er_inner, textvariable=self._ver,
                     font=ctk.CTkFont(size=22, weight="bold"),
                     text_color=C["green"]).pack(side="left", padx=(16, 0))

        self._ver_hint = ctk.StringVar(value="")
        ctk.CTkLabel(er_inner, textvariable=self._ver_hint,
                     font=ctk.CTkFont(size=10),
                     text_color=C["txt4"]).pack(side="left", padx=(8, 0))

        self._vlast_post = ctk.StringVar(value="—")
        ctk.CTkLabel(er_inner, text="Last post:",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=C["txt4"]).pack(side="right", padx=(0, 4))
        ctk.CTkLabel(er_inner, textvariable=self._vlast_post,
                     font=ctk.CTkFont(size=11),
                     text_color=C["txt2"]).pack(side="right", padx=(0, 18))

        # ── BIO + INFO  2-column grid ─────────────────────────────────────────
        bio_row = ctk.CTkFrame(sc, fg_color="transparent")
        bio_row.pack(fill="x", padx=4, pady=(10, 4))
        bio_row.columnconfigure(0, weight=3)
        bio_row.columnconfigure(1, weight=2)

        # Bio card with gradient left border (4px colored frame)
        bio_card = ctk.CTkFrame(bio_row, fg_color=C["card"],
                                corner_radius=14,
                                border_color=C["border"], border_width=1)
        bio_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        # Accent bar left side
        bio_hdr = ctk.CTkFrame(bio_card, fg_color="transparent")
        bio_hdr.pack(fill="x", padx=14, pady=(12, 4))
        ctk.CTkFrame(bio_hdr, width=4, fg_color=C["cyan"],
                     corner_radius=2).pack(side="left", fill="y",
                                           padx=(0, 8), pady=2)
        ctk.CTkLabel(bio_hdr, text="Biography",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C["cyan"]).pack(side="left")

        self._bio_tb = ctk.CTkTextbox(
            bio_card, height=110,
            fg_color=C["bg3"], text_color=C["txt2"],
            font=ctk.CTkFont(size=11), wrap="word",
            corner_radius=8,
            border_color=C["border"], border_width=1)
        self._bio_tb.pack(fill="x", padx=14, pady=(0, 14))
        _tb_set(self._bio_tb, "No biography.")

        # Quick-info card (right column)
        qi_card = ctk.CTkFrame(bio_row, fg_color=C["card"],
                               corner_radius=14,
                               border_color=C["border"], border_width=1)
        qi_card.grid(row=0, column=1, sticky="nsew")

        qi_hdr = ctk.CTkFrame(qi_card, fg_color="transparent")
        qi_hdr.pack(fill="x", padx=14, pady=(12, 4))
        ctk.CTkFrame(qi_hdr, width=4, fg_color=C["purple"],
                     corner_radius=2).pack(side="left", fill="y",
                                           padx=(0, 8), pady=2)
        ctk.CTkLabel(qi_hdr, text="Quick Info",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=C["purple"]).pack(side="left")

        ctk.CTkFrame(qi_card, height=1,
                     fg_color=C["border"]).pack(fill="x", padx=14, pady=(0, 8))

        self._qi_vars = {}
        for k, label, col in (
            ("📷 Posts",    "Posts",        C["txt"]),
            ("📅 Joined",   "Account Age",  C["txt2"]),
            ("⏰ Last Post","Last Active",  C["yellow"]),
            ("🔒 Privacy",  "Privacy",      C["txt2"]),
        ):
            r = ctk.CTkFrame(qi_card, fg_color="transparent")
            r.pack(fill="x", padx=14, pady=3)
            ctk.CTkLabel(r, text=k,
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color=C["txt4"], width=110, anchor="w").pack(side="left")
            v = ctk.StringVar(value="—")
            ctk.CTkLabel(r, textvariable=v,
                         font=ctk.CTkFont(size=11),
                         text_color=col, anchor="w").pack(side="left", fill="x", expand=True)
            self._qi_vars[k] = v
        ctk.CTkFrame(qi_card, height=8, fg_color="transparent").pack()

    def _draw_hero_bg(self, event=None):
        """Draw animated gradient on the hero canvas."""
        cv = self._hero_canvas
        try:
            cv.update_idletasks()
            w = cv.winfo_width()
            h = cv.winfo_height()
            if w < 10: w = 900
            if h < 10: h = 190
        except:
            w, h = 900, 190
        
        # Create PIL image
        banner = make_hero_banner(w, h)
        
        # Convert to PhotoImage for tkinter Canvas
        self._hero_photo = ImageTk.PhotoImage(banner)
        
        cv.delete("all")
        cv.create_image(0, 0, anchor="nw", image=self._hero_photo)
        
        # Watermark text - use RGB hex without alpha
        cv.create_text(w-16, h-10,
                       text=f"InstaGrep v4.0  ⚡  Instagram OSINT",
                       fill="#222222", anchor="se",
                       font=("Courier New", 9))
        
        # Scan line effect - use RGB without alpha
        for y in range(0, h, 22):
            cv.create_line(0, y, w, y, fill="#1a1a2a", width=1)

    def _update_stat_sparklines(self, medias):
        """Draw mini sparklines on stat cards using last 12 posts engagement."""
        if not medias: return
        recent = medias[:12]
        likes  = [getattr(m,"like_count",0) or 0 for m in recent][::-1]
        for key, color, _ in (
            ("followers", C["cyan"],   C["cyan_dim"]),
            ("following", C["purple"], C["purple_dim"]),
            ("posts",     C["green"],  C["green_dim"]),
        ):
            lbl = self._stat_sparkline.get(key)
            if not lbl: continue
            spark_img = make_sparkline(likes, w=110, h=26, color=color)
            ci = ctk.CTkImage(spark_img, spark_img, size=(110, 26))
            lbl.configure(image=ci, text="")
            lbl._img = ci

    def _render_placeholder_avatar(self):
        ph   = make_placeholder_avatar(148)
        ring = make_avatar_ring(ph, 148)
        ci   = ctk.CTkImage(ring, ring, size=(182, 182))
        self._pic_lbl.configure(image=ci, text="")
        self._pic_lbl._img = ci

    def _render_badges(self, is_verified, is_private, is_business):
        for w in self._badge_frame.winfo_children():
            w.destroy()
        badges = []
        if is_verified: badges.append(("✅ Verified",  C["green"],  C["green_dim"]))
        if is_business: badges.append(("🏢 Business",  C["orange"], C["card3"]))
        if is_private:  badges.append(("🔒 Private",   C["yellow"], C["card3"]))
        else:           badges.append(("🌐 Public",    C["cyan"],   C["cyan_dim"]))
        for txt, fg, bg in badges:
            pill = ctk.CTkLabel(self._badge_frame, text=txt,
                                font=ctk.CTkFont(size=10, weight="bold"),
                                text_color=fg, fg_color=bg,
                                corner_radius=7, padx=9, pady=4)
            pill.pack(side="left", padx=(0, 5), pady=2)
            # Hover
            pill.bind("<Enter>", lambda e,p=pill,f=fg: p.configure(fg_color=f, text_color=C["bg"]))
            pill.bind("<Leave>", lambda e,p=pill,f=fg,b=bg: p.configure(fg_color=b, text_color=f))

    def _animate_stat(self, key, target_raw: int, duration_ms=900):
        """Ease-out count-up with colour pulse on finish."""
        lbl, color = self._stat_labels.get(key, (None, None))
        if lbl is None: return
        steps   = 36
        step_ms = max(duration_ms // steps, 12)

        if key in self._count_jobs:
            try: self.after_cancel(self._count_jobs[key])
            except: pass
        lbl.configure(text="0", text_color=color)

        def _tick(step=0):
            pct  = step / steps
            ease = 1 - (1 - pct) ** 3      # ease-out cubic
            curr = int(target_raw * ease)
            lbl.configure(text=fmt_num(curr))
            if step < steps:
                self._count_jobs[key] = self.after(step_ms, _tick, step + 1)
            else:
                lbl.configure(text=fmt_num(target_raw))
                # Brief white flash on finish
                lbl.configure(text_color="#ffffff")
                self.after(120, lambda: lbl.configure(text_color=color))

        self.after(0, _tick)

    # ══════════════════════════════════════════════════════════════════════════
    #  Contact tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_contact_tab(self):
        t  = self._tabs.tab("📬  Contact")
        sc = ctk.CTkScrollableFrame(t, fg_color="transparent")
        sc.pack(fill="both", expand=True)

        biz = self._card(sc, "Business Contact", C["orange"])
        biz.pack(fill="x", padx=4, pady=(4,8))
        self._vemail = ctk.StringVar(value="—")
        self._vphone = ctk.StringVar(value="—")
        self._vaddr  = ctk.StringVar(value="—")
        for lbl, var in (("Email:", self._vemail),
                         ("Phone:", self._vphone),
                         ("Address:", self._vaddr)):
            self._irow(biz, lbl, var)
        ctk.CTkFrame(biz, height=8, fg_color="transparent").pack()

        row2 = ctk.CTkFrame(sc, fg_color="transparent")
        row2.pack(fill="x", padx=4)
        em = self._card(row2, "Emails in Captions", C["cyan"])
        em.pack(side="left", fill="both", expand=True, padx=(0,6))
        self._emails_tb = self._textbox(em, 140)
        self._emails_tb.pack(fill="x", padx=14, pady=(0,14))
        _tb_set(self._emails_tb, "Run analysis first.")

        ph = self._card(row2, "Phone Numbers in Captions", C["green"])
        ph.pack(side="left", fill="both", expand=True)
        self._phones_tb = self._textbox(ph, 140)
        self._phones_tb.pack(fill="x", padx=14, pady=(0,14))
        _tb_set(self._phones_tb, "Run analysis first.")

        bio_i = self._card(sc, "Bio Intelligence", C["purple"])
        bio_i.pack(fill="x", padx=4, pady=(8,0))
        self._bio_intel_tb = self._textbox(bio_i, 130)
        self._bio_intel_tb.pack(fill="x", padx=14, pady=(0,14))
        _tb_set(self._bio_intel_tb, "Run analysis first.")

    # ══════════════════════════════════════════════════════════════════════════
    #  Locations tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_locations_tab(self):
        t  = self._tabs.tab("📍  Locations")
        sc = ctk.CTkScrollableFrame(t, fg_color="transparent")
        sc.pack(fill="both", expand=True)
        card = self._card(sc, "Location Intelligence — Last 50 Posts", C["green"])
        card.pack(fill="x", padx=4, pady=4)
        self._loc_tb = self._textbox(card, 560, mono=True, color=C["green"])
        self._loc_tb.pack(fill="x", padx=14, pady=(0,14))
        _tb_set(self._loc_tb, "Run analysis first.")

    # ══════════════════════════════════════════════════════════════════════════
    #  Posts tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_posts_tab(self):
        t = self._tabs.tab("🗂  Posts")
        sf = ctk.CTkFrame(t, fg_color=C["card"], height=46, corner_radius=8)
        sf.pack(fill="x", padx=5, pady=(5,4))
        sf.pack_propagate(False)
        self._posts_stat = ctk.StringVar(value="Run analysis to view posts")
        ctk.CTkLabel(sf, textvariable=self._posts_stat,
                     font=ctk.CTkFont(size=12),
                     text_color=C["txt2"]).pack(side="left", padx=14)
        self._posts_sc = ctk.CTkScrollableFrame(t, fg_color="transparent")
        self._posts_sc.pack(fill="both", expand=True, padx=5)
        ctk.CTkLabel(self._posts_sc,
                     text="No posts yet — run analysis",
                     font=ctk.CTkFont(size=14),
                     text_color=C["txt4"]).pack(pady=50)

    # ══════════════════════════════════════════════════════════════════════════
    #  Activity tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_activity_tab(self):
        t    = self._tabs.tab("📊  Activity")
        main = ctk.CTkFrame(t, fg_color="transparent")
        main.pack(fill="both", expand=True)

        lp = ctk.CTkFrame(main, fg_color="transparent", width=370)
        lp.pack(side="left", fill="y", padx=(4,6), pady=4)
        lp.pack_propagate(False)

        act_c = self._card(lp, "Last Activity", C["yellow"])
        act_c.pack(fill="x", pady=(0,8))
        self._vlast_story  = ctk.StringVar(value="—")
        self._vlive        = ctk.StringVar(value="—")
        self._vstories_act = ctk.StringVar(value="—")
        self._vlive_link   = ctk.StringVar(value="—")
        for lbl,var in (("📷  Last Post:","_vlast_post"),
                        ("📖  Last Story:","_vlast_story"),
                        ("🔴  Live?","_vlive"),
                        ("💫  Stories Active?","_vstories_act")):
            self._irow(act_c, lbl, getattr(self, var), lbl_w=160)
        ctk.CTkFrame(act_c, height=8, fg_color="transparent").pack()

        met_c = self._card(lp, "Engagement Metrics", C["cyan"])
        met_c.pack(fill="x", pady=(0,8))
        self._vavgposts  = ctk.StringVar(value="—")
        self._vpeak_hr   = ctk.StringVar(value="—")
        self._vpeak_day  = ctk.StringVar(value="—")
        self._vavg_likes = ctk.StringVar(value="—")
        self._vavg_cmts  = ctk.StringVar(value="—")
        for lbl,var in (("Avg Posts/Day:",self._vavgposts),
                        ("Peak Hour:",self._vpeak_hr),
                        ("Peak Day:",self._vpeak_day),
                        ("Avg Likes:",self._vavg_likes),
                        ("Avg Comments:",self._vavg_cmts)):
            self._irow(met_c, lbl, var, lbl_w=150)

        ctk.CTkFrame(met_c, height=5, fg_color="transparent").pack()
        ctk.CTkLabel(met_c, text="Top Hashtags:",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["txt4"]).pack(anchor="w", padx=15)
        self._tags_tb = self._textbox(met_c, 100, color=C["cyan"])
        self._tags_tb.pack(fill="x", padx=15, pady=(4,15))
        _tb_set(self._tags_tb, "—")

        rp = ctk.CTkScrollableFrame(main, fg_color="transparent")
        rp.pack(side="left", fill="both", expand=True, pady=4)

        self._chart_frames = {}
        self._chart_phs    = {}
        for key, ttl, col in (
            ("hours","🕐  Most Active Hours",  C["cyan"]),
            ("days", "📅  Most Active Days",   C["purple"]),
            ("tags", "🏷  Top Hashtags",        C["green"]),
            ("eng",  "📈  Engagement Over Time",C["green"]),
        ):
            f  = self._card(rp)
            f.pack(fill="x", pady=(0,8))
            ph = ctk.CTkLabel(f, text=f"{ttl} — chart will appear after analysis",
                              text_color=C["txt4"], height=200)
            ph.pack()
            self._chart_frames[key] = f
            self._chart_phs[key]    = ph

    # ══════════════════════════════════════════════════════════════════════════
    #  Downloads tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_downloads_tab(self):
        t    = self._tabs.tab("⬇  Downloads")
        ctrl = self._card(t, "Download Media", C["green"])
        ctrl.pack(fill="x", padx=5, pady=(5,8))

        row = ctk.CTkFrame(ctrl, fg_color="transparent")
        row.pack(fill="x", padx=14, pady=(0,12))
        self._dl_pic_var  = ctk.BooleanVar(value=True)
        self._dl_post_var = ctk.BooleanVar(value=True)
        for txt, var in (("Profile Picture",self._dl_pic_var),
                         ("Posts (images/videos)",self._dl_post_var)):
            ctk.CTkCheckBox(row, text=txt, variable=var,
                            font=ctk.CTkFont(size=12),
                            text_color=C["txt2"],
                            fg_color=C["blue"],
                            hover_color=C["blue2"]).pack(side="left",padx=(0,20))

        btns = ctk.CTkFrame(ctrl, fg_color="transparent")
        btns.pack(fill="x", padx=14, pady=(0,14))
        self._dl_btn = ctk.CTkButton(
            btns, text="⬇  Start Download", command=self._dl_start,
            width=175, height=38,
            font=ctk.CTkFont(size=13,weight="bold"),
            fg_color=C["green2"], hover_color=C["green"],
            text_color=C["bg"], corner_radius=8)
        self._dl_btn.pack(side="left")
        self._dl_cancel = ctk.CTkButton(
            btns, text="✕  Cancel", command=self._dl_cancel,
            width=100, height=38,
            font=ctk.CTkFont(size=12),
            fg_color=C["card"], hover_color=C["red2"],
            text_color=C["txt2"], state="disabled", corner_radius=8)
        self._dl_cancel.pack(side="left", padx=8)

        prog_c = self._card(t)
        prog_c.pack(fill="x", padx=5, pady=(0,8))
        self._dl_bar = ctk.CTkProgressBar(prog_c, mode="determinate",
                                          fg_color=C["bg3"],
                                          progress_color=C["green2"])
        self._dl_bar.pack(fill="x", padx=14, pady=(8,4))
        self._dl_bar.set(0)
        self._dl_sv = ctk.StringVar(value="No downloads yet.")
        ctk.CTkLabel(prog_c, textvariable=self._dl_sv,
                     font=ctk.CTkFont(size=11),
                     text_color=C["txt2"]).pack(anchor="w",padx=14,pady=(0,10))

        log_c = self._card(t, "Download Log", C["txt4"])
        log_c.pack(fill="both", expand=True, padx=5)
        self._dl_log = self._textbox(log_c, 300, mono=True, color=C["green"])
        self._dl_log.pack(fill="both", expand=True, padx=14, pady=(0,14))
        _tb_set(self._dl_log, "Download log …")

    # ══════════════════════════════════════════════════════════════════════════
    #  Export tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_export_tab(self):
        t  = self._tabs.tab("📤  Export")
        sc = ctk.CTkScrollableFrame(t, fg_color="transparent")
        sc.pack(fill="both", expand=True)
        card = self._card(sc, "Export Reports", C["blue"])
        card.pack(fill="x", padx=5, pady=5)

        for txt, desc, col, cmd in (
            ("💾  JSON",   "Raw developer data",                    C["blue"],   self._save_json),
            ("📊  Excel",  "Multi-sheet workbook",                   C["green2"], self._export_excel),
            ("🌐  HTML",   "Interactive HTML dashboard",             C["cyan"],   self._export_html),
            ("📄  PDF",    "Professional report with charts",        C["purple"], self._export_pdf),
        ):
            r = ctk.CTkFrame(card, fg_color=C["card2"], corner_radius=10)
            r.pack(fill="x", padx=14, pady=5)
            ctk.CTkLabel(r, text=txt,
                         font=ctk.CTkFont(size=13,weight="bold"),
                         text_color=col).pack(side="left", padx=14, pady=12)
            ctk.CTkLabel(r, text=desc,
                         font=ctk.CTkFont(size=11),
                         text_color=C["txt4"]).pack(side="left")
            ctk.CTkButton(r, text="Export →", command=cmd,
                          width=100, height=30,
                          font=ctk.CTkFont(size=11),
                          fg_color=C["blue2"], hover_color=C["blue"],
                          text_color=C["txt"]).pack(side="right", padx=14)
        ctk.CTkFrame(card, height=10, fg_color="transparent").pack()

        log_c = self._card(sc, "Export Log", C["txt4"])
        log_c.pack(fill="x", padx=5, pady=(8,5))
        self._export_log = self._textbox(log_c, 180, mono=True)
        self._export_log.pack(fill="x", padx=14, pady=(0,14))
        _tb_set(self._export_log, "Export log …")

    # ══════════════════════════════════════════════════════════════════════════
    #  Raw tab
    # ══════════════════════════════════════════════════════════════════════════

    def _build_raw_tab(self):
        t = self._tabs.tab("{}  Raw")
        tb_bar = ctk.CTkFrame(t, fg_color=C["card"], height=44, corner_radius=8)
        tb_bar.pack(fill="x", padx=5, pady=(5,4)); tb_bar.pack_propagate(False)
        ctk.CTkLabel(tb_bar, text="Raw JSON Data",
                     font=ctk.CTkFont(size=13,weight="bold"),
                     text_color=C["txt2"]).pack(side="left", padx=14)
        for txt, cmd in (("💾  Save",self._save_json),("📋  Copy",self._copy_json)):
            ctk.CTkButton(tb_bar, text=txt, command=cmd,
                          width=110, height=32,
                          fg_color=C["bg3"], hover_color=C["card2"],
                          text_color=C["txt"],
                          font=ctk.CTkFont(size=12)).pack(side="right", padx=5)
        self._raw_tb = ctk.CTkTextbox(
            t, fg_color=C["bg3"], text_color=C["green"],
            font=ctk.CTkFont(family="Courier New",size=11), wrap="none")
        self._raw_tb.pack(fill="both", expand=True, padx=5, pady=(0,5))
        _tb_set(self._raw_tb, '{\n  "status": "No data yet"\n}')

    # ══════════════════════════════════════════════════════════════════════════
    #  Widget factory helpers (instance-level)
    # ══════════════════════════════════════════════════════════════════════════

    def _card(self, parent, title=None, accent=None, **kw):
        f = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=14, **kw)
        if title:
            hdr = ctk.CTkFrame(f, fg_color="transparent")
            hdr.pack(fill="x", padx=16, pady=(13,0))
            if accent:
                ctk.CTkLabel(hdr, text="▌",
                             font=ctk.CTkFont(size=18,weight="bold"),
                             text_color=accent).pack(side="left", padx=(0,6))
            ctk.CTkLabel(hdr, text=title,
                         font=ctk.CTkFont(size=13,weight="bold"),
                         text_color=C["cyan"]).pack(side="left")
            ctk.CTkFrame(f, height=1, fg_color=C["border"]).pack(
                fill="x", padx=16, pady=(6,8))
        return f

    def _irow(self, parent, label, var=None, lbl_w=165, accent=None):
        r = ctk.CTkFrame(parent, fg_color="transparent")
        r.pack(fill="x", padx=16, pady=3)
        ctk.CTkLabel(r, text=label,
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["txt4"], width=lbl_w, anchor="w").pack(side="left")
        kw = dict(textvariable=var) if var else dict(text="—")
        lbl = ctk.CTkLabel(r, **kw,
                           font=ctk.CTkFont(size=11),
                           text_color=accent or C["txt"],
                           anchor="w", wraplength=530, justify="left")
        lbl.pack(side="left", fill="x", expand=True)
        return lbl

    def _textbox(self, parent, height=120, mono=False, color=None):
        return ctk.CTkTextbox(parent, height=height,
                              fg_color=C["bg3"],
                              text_color=color or C["txt"],
                              font=ctk.CTkFont(
                                  family="Courier New" if mono else "default",
                                  size=11),
                              wrap="word", corner_radius=8)

    # ══════════════════════════════════════════════════════════════════════════
    #  Queue poll loop
    # ══════════════════════════════════════════════════════════════════════════

    def _poll(self):
        try:
            while True:
                kind, payload = self._q.get_nowait()
                handler = {
                    "status":           lambda p: self._typewriter(p),
                    "profile":          self._on_profile,
                    "medias":           self._on_medias,
                    "stories":          self._on_stories,
                    "live":             self._on_live,
                    "raw":              self._on_raw,
                    "charts":           self._on_charts,
                    "pic":              self._on_pic,
                    "error":            self._on_error,
                    "done":             self._on_done,
                    "cookie_ok":        self._on_cookie_ok,
                    "cookie_none":      self._on_cookie_none,
                    "cookie_err":       self._on_cookie_err,
                    "_deferred_start":  self._launch_worker,
                    "dl_progress":      lambda p: (self._dl_bar.set(p[0]/max(p[1],1)),
                                                   self._dl_sv.set(f"Downloading  {p[0]}/{p[1]}  ({p[0]/max(p[1],1)*100:.0f}%)")),
                    "dl_log":           lambda p: self._append_log(self._dl_log, p),
                    "dl_done":          self._on_dl_done,
                    "export_log":       lambda p: self._append_log(self._export_log, p),
                    "toast":            lambda p: Toast.show(self, p[0], p[1]),
                }.get(kind, lambda p: None)
                handler(payload)
        except queue.Empty:
            pass
        self.after(80, self._poll)

    def _put(self, kind, payload=None):
        self._q.put((kind, payload))

    def _append_log(self, tb, msg):
        tb.configure(state="normal")
        tb.insert("end", msg + "\n")
        tb.see("end")
        tb.configure(state="disabled")

    # ══════════════════════════════════════════════════════════════════════════
    #  Typewriter animation
    # ══════════════════════════════════════════════════════════════════════════

    def _typewriter(self, msg, speed=16):
        if self._tw_job:
            try: self.after_cancel(self._tw_job)
            except: pass
        self._tw_target  = msg
        self._tw_current = ""
        self._sv.set("")
        self._tw_tick(speed)

    def _tw_tick(self, speed):
        if len(self._tw_current) < len(self._tw_target):
            self._tw_current = self._tw_target[:len(self._tw_current)+1]
            self._sv.set(self._tw_current)
            self._tw_job = self.after(speed, self._tw_tick, speed)

    # ══════════════════════════════════════════════════════════════════════════
    #  Session management
    # ══════════════════════════════════════════════════════════════════════════

    def _activate_saved(self):
        c = self._sessions.get_active()
        if c:
            self._put("cookie_ok", (c, f"Saved ({self._sessions.active_name})"))
        self._refresh_session_menu()

    def _refresh_session_menu(self):
        ns = self._sessions.list()
        self._sess_menu.configure(values=ns or ["—"])
        if self._sessions.active_name:
            self._sess_var.set(self._sessions.active_name)

    def _switch_session(self, name):
        if name in self._sessions.list():
            self._sessions.set_active(name)
            c = self._sessions.get_active()
            if c: self._put("cookie_ok", (c, f"Saved ({name})"))

    # ══════════════════════════════════════════════════════════════════════════
    #  Cookie engine
    # ══════════════════════════════════════════════════════════════════════════

    def _import_cookie(self):
        path = filedialog.askopenfilename(
            title="Select cookies.json",
            filetypes=[("JSON","*.json"),("All","*.*")])
        if path:
            threading.Thread(target=self._load_cookie_file,
                             args=(path,), daemon=True).start()

    def _load_cookie_file(self, path):
        self._ck_dot.configure(text_color=C["s_warn"])
        c = self._parse_cookie_file(path)
        if c:
            name = Path(path).stem + f"_{datetime.now():%H%M%S}"
            self._sessions.add(name, c)
            self._put("cookie_ok", (c, f"File: {Path(path).name}"))
        else:
            self._put("cookie_err", f"No sessionid in {Path(path).name}")

    @staticmethod
    def _parse_cookie_file(filename) -> dict | None:
        try:
            data = json.loads(Path(filename).read_text(encoding="utf-8"))
            cd   = {}
            if isinstance(data, list):
                for item in data:
                    n = item.get("name") or item.get("Name") or item.get("key","")
                    v = item.get("value") or item.get("Value","")
                    if n: cd[n] = v
            elif isinstance(data, dict):
                if "cookies" in data:
                    for item in data["cookies"]:
                        n=item.get("name",""); v=item.get("value","")
                        if n: cd[n]=v
                else:
                    cd = data
            return cd if any("sessionid" in k.lower() for k in cd) else None
        except: return None

    def _probe_async(self):
        self._ck_dot.configure(text_color=C["s_warn"])
        self._ck_sv.set("Scanning …")
        threading.Thread(target=self._probe_worker, daemon=True).start()

    def _probe_worker(self):
        for fname in ("cookies.json","instagram_cookies.json"):
            c = self._parse_cookie_file(fname)
            if c:
                self._sessions.add(f"file_{fname}", c)
                self._put("cookie_ok", (c, f"File: {fname}")); return
        if not COOKIES_OK:
            self._put("cookie_err","browser-cookie3 not installed"); return
        for bname, loader in BROWSER_LOADERS:
            try:
                c = {c.name:c.value for c in loader(domain_name=".instagram.com")}
                if "sessionid" in c:
                    self._sessions.add(bname, c)
                    self._put("cookie_ok", (c, bname)); return
            except: pass
        self._put("cookie_none","No Instagram session found")

    @staticmethod
    def _build_client(cookies) -> "Client":
        cl = Client(); cl.delay_range=[1,3]; cl.set_settings({})
        for n,v in cookies.items():
            cl.private.cookies.set(n,v,domain=".instagram.com")
            cl.public.cookies.set(n,v,domain=".instagram.com")
        if "csrftoken" in cookies:
            cl.private.headers.update({"X-CSRFToken":cookies["csrftoken"]})
        return cl

    # ══════════════════════════════════════════════════════════════════════════
    #  Cookie event handlers
    # ══════════════════════════════════════════════════════════════════════════

    def _on_cookie_ok(self, payload):
        cookies, bname = payload
        self._cl = self._build_client(cookies)
        uid = cookies.get("ds_user_id","")
        self._ck_dot.configure(text_color=C["s_ok"])
        self._ck_sv.set(f"✅  {bname}" + (f"  │  uid:{uid}" if uid else "") + "  —  ready")
        self._dot.configure(text_color=C["s_ok"])
        self._typewriter(f"✅  Session loaded from {bname}  —  enter a target username and press ANALYZE")
        self._refresh_session_menu()
        Toast.show(self, f"Session loaded: {bname}", "ok")

    def _on_cookie_none(self, detail):
        self._cl = None
        self._ck_dot.configure(text_color=C["s_err"])
        self._ck_sv.set("❌  No session — import cookies.json")
        self._typewriter(f"No session: {detail}")
        Toast.show(self, "No Instagram session found", "warn")

    def _on_cookie_err(self, msg):
        self._cl = None
        self._ck_dot.configure(text_color=C["s_warn"])
        self._ck_sv.set(f"⚠  {msg}")
        self._typewriter(f"Session error: {msg}")

    # ══════════════════════════════════════════════════════════════════════════
    #  Analysis
    # ══════════════════════════════════════════════════════════════════════════

    def _start(self):
        if self._analyzing: return
        self._close_recent()
        u = self._uv.get().strip().lstrip("@")
        if not u:
            Toast.show(self, "Enter a username first", "warn"); return
        if self._cl is None:
            self._typewriter("⏳  No session — scanning …")
            def _probe_then():
                self._probe_worker()
                time.sleep(0.6)
                if self._cl is not None:
                    self._put("_deferred_start", u)
                else:
                    self._put("toast", ("No session found — import cookies.json", "err"))
            threading.Thread(target=_probe_then, daemon=True).start()
            return
        self._launch_worker(u)

    def _launch_worker(self, username):
        self._analyzing = True
        self._btn.configure(state="disabled", text="⏳  Analyzing …")
        self._prog.pack(side="right", padx=10)
        self._prog.start()
        threading.Thread(target=self._worker, args=(username,), daemon=True).start()

    def _worker(self, raw: str):
        u = raw.strip().lower()
        if "instagram.com/" in u: u=u.split("instagram.com/")[-1]
        u = u.lstrip("@").rstrip("/").split("?")[0].strip()

        cl = self._cl
        if cl is None:
            self._put("error","No session"); return

        def st(m): self._put("status", m)

        try:
            st(f"Fetching @{u} …")
            try:
                user = retry(cl.user_info_by_username_v1, u)
            except LoginRequired:
                self._cl = None
                self._put("cookie_none","Session expired")
                self._put("error","Session expired — re-import cookies.json"); return
            except Exception:
                uid  = retry(cl.user_id_from_username, u)
                user = retry(cl.user_info, uid)

            self._put("profile", user)
            time.sleep(0.5)

            st("Fetching posts …")
            medias = []
            try:
                medias = retry(cl.user_medias, user.pk, 50)
            except LoginRequired:
                self._cl = None
                self._put("error","Session expired mid-analysis"); return
            except Exception as e:
                st(f"Posts limited — {_short(e,60)}")
            self._put("medias", medias)
            self._medias_cache = medias
            time.sleep(0.6)

            st("Checking stories …")
            stories = []
            try: stories = retry(cl.user_stories, user.pk, attempts=2)
            except: pass
            self._put("stories", stories)
            time.sleep(0.3)

            st("Checking live …")
            live = {"is_live": bool(getattr(user,"is_live",False)), "link": None}
            try:
                bcs = retry(cl.user_broadcasts, user.pk, attempts=2) or []
                if bcs:
                    live["is_live"]=True
                    live["link"]=f"https://www.instagram.com/{u}/live/"
            except: pass
            self._put("live", live)

            act = self._compute_activity(medias)
            contact = self._extract_contact_data(user, medias)
            raw_data = {
                "username":      u,
                "analyzed_at":   datetime.now().isoformat(),
                "session":       self._sessions.active_name or "unknown",
                "profile":       self._ser(user.dict() if hasattr(user,"dict") else {}),
                "posts_fetched": len(medias),
                "stories_active":len(stories),
                "contact":       contact,
                "activity":      act,
                "posts":         [self._ser_media(m) for m in medias],
            }
            self._put("raw", raw_data)

            # Save to recent searches
            fn = str(getattr(user,"full_name","") or "")
            self._recent.add(u, fn)

            self._put("done", f"✅  @{u}  —  {len(medias)} posts  │  {len(stories)} stories")

        except UserNotFound:
            self._put("error", f"@{u} not found")
        except PrivateAccount:
            self._put("error", f"@{u} is private")
        except Exception as e:
            self._put("error", _short(e))

    # ══════════════════════════════════════════════════════════════════════════
    #  Queue event handlers
    # ══════════════════════════════════════════════════════════════════════════

    def _on_profile(self, user):
        # Profile picture
        pic_url = str(getattr(user,"profile_pic_url_hd","") or
                      getattr(user,"profile_pic_url","") or "")
        if pic_url and pic_url != "None":
            threading.Thread(target=self._fetch_pic,
                             args=(pic_url,), daemon=True).start()

        def g(a, fb="—"):
            v = getattr(user, a, None)
            return str(v) if v else fb

        # Hero identity
        self._lbl_fullname.configure(text=g("full_name", "—"))
        self._lbl_username.configure(text=f"@{g('username','—')}")
        self._lbl_uid.configure(text=f"  UID: {g('pk')}")

        priv  = getattr(user,"is_private",  False)
        biz   = getattr(user,"is_business", False)
        verif = getattr(user,"is_verified", False)
        self._vtype.set("🏢  Business" if biz else "🔒  Private" if priv else "🌐  Public")
        self._vverif.set("✅  Yes" if verif else "No")
        self._vurl.set(g("external_url"))
        self._vcat.set(str(getattr(user,"category_name","") or
                          getattr(user,"category","") or "—"))

        _tb_set(self._bio_tb, str(getattr(user,"biography","") or "") or "No biography.")
        self._render_badges(verif, priv, biz)

        # Stat cards with count-up animation
        fc = getattr(user,"follower_count",0) or 0
        fg = getattr(user,"following_count",0) or 0
        mc = getattr(user,"media_count",0) or 0
        self._animate_stat("followers", fc)
        self._animate_stat("following", fg, duration_ms=650)
        self._animate_stat("posts",     mc, duration_ms=450)
        # Store for engagement rate calc
        self._cached_follower_count = fc

        # Quick info card
        mc_str   = fmt_num(mc)
        priv_str = "🔒 Private" if priv else "🌐 Public"
        self._qi_vars.get("📷 Posts",    ctk.StringVar()).set(mc_str)
        self._qi_vars.get("🔒 Privacy",  ctk.StringVar()).set(priv_str)

        # Contact tab
        email = str(getattr(user,"business_email","") or getattr(user,"public_email","") or "—")
        phone = str(getattr(user,"business_phone_number","") or
                    getattr(user,"contact_phone_number","") or "—")
        self._vemail.set(email if email not in ("","None") else "—")
        self._vphone.set(phone if phone not in ("","None") else "—")
        addr_raw = getattr(user,"business_address_json",None)
        if isinstance(addr_raw,dict):
            addr = ", ".join(str(addr_raw.get(k,"") or "") for k in
                             ("street_address","city_name","zip_code","country_code"))
            self._vaddr.set(addr or "—")

    def _on_medias(self, medias):
        self._render_posts(medias)
        self._render_locations(medias)
        bio = self._bio_tb.get("1.0","end").strip()
        self._render_bio_intel(bio, medias)

        act = self._compute_activity(medias)
        self._apply_activity_ui(act, medias)

        # Update quick-info last-active
        lp = act.get("last_post","—")
        if lp != "—":
            self._vlast_post.set(lp)
            self._qi_vars.get("⏰ Last Post", ctk.StringVar()).set(lp[:10])

        # Sparklines on stat cards
        self._update_stat_sparklines(medias)

        # Engagement rate
        fc = getattr(self, "_cached_follower_count", 0)
        if fc > 0 and medias:
            tl = sum(getattr(m,"like_count",0) or 0 for m in medias)
            tc = sum(getattr(m,"comment_count",0) or 0 for m in medias)
            er = ((tl+tc)/len(medias)) / fc * 100
            self._ver.set(f"{er:.2f}%")
            grade = "🔥 Excellent" if er>6 else "✅ Good" if er>3 else "⚠ Average" if er>1 else "❄ Low"
            self._ver_hint.set(f"({grade})")
        else:
            self._ver.set("N/A")

    def _on_stories(self, stories):
        if stories:
            ta = getattr(stories[0],"taken_at",None)
            self._vlast_story.set(ta.strftime("%Y-%m-%d  %H:%M") if ta else "—")
            self._vstories_act.set(f"YES  ({len(stories)})")
        else:
            self._vlast_story.set("None / private")
            self._vstories_act.set("NO")

    def _on_live(self, live):
        if live.get("is_live"):
            self._vlive.set("🔴  LIVE NOW!")
        else:
            self._vlive.set("—")

    def _on_raw(self, data):
        self._data = data
        _tb_set(self._raw_tb, json.dumps(data, indent=2, default=str))

    def _on_charts(self, charts: dict):
        for key, img in charts.items():
            if img is None: continue
            frame = self._chart_frames.get(key)
            if not frame: continue
            ph = self._chart_phs.get(key)
            if ph:
                try: ph.destroy(); self._chart_phs[key] = None
                except: pass
            frame.update_idletasks()
            fw  = max(frame.winfo_width() - 24, 620)
            # Preserve chart aspect ratio (10:3.2 figsize)
            fh  = int(fw * 3.2 / 10.0)
            ci  = ctk.CTkImage(img, img, size=(fw, fh))
            lbl = ctk.CTkLabel(frame, image=ci, text="")
            lbl._img = ci
            lbl.pack(padx=10, pady=(4, 12))

    def _on_pic(self, ci):
        self._pic_lbl.configure(image=ci)
        self._pic_lbl._img = ci

    def _on_error(self, msg):
        self._sv.set(f"❌  {msg}")
        self._dot.configure(text_color=C["s_err"])
        self._prog.stop(); self._prog.pack_forget()
        self._btn.configure(state="normal", text="⚡  ANALYZE")
        self._analyzing = False
        Toast.show(self, f"Error: {msg[:60]}", "err")

    def _on_done(self, msg):
        self._typewriter(msg)
        self._dot.configure(text_color=C["s_ok"])
        self._prog.stop(); self._prog.pack_forget()
        self._btn.configure(state="normal", text="⚡  ANALYZE")
        self._analyzing = False
        Toast.show(self, "Analysis complete ✅", "ok")

    def _on_dl_done(self, stats):
        self._dl_btn.configure(state="normal", text="⬇  Start Download")
        self._dl_cancel.configure(state="disabled")
        self._append_log(self._dl_log,
                         f"\n✅  Done — {stats['downloaded']} downloaded, "
                         f"{stats['failed']} failed")
        Toast.show(self, f"Download done: {stats['downloaded']} files", "ok")

    # ══════════════════════════════════════════════════════════════════════════
    #  Data helpers
    # ══════════════════════════════════════════════════════════════════════════

    def _fetch_pic(self, url):
        try:
            r    = requests.get(url, timeout=12, headers={"User-Agent":"Mozilla/5.0"})
            raw  = Image.open(io.BytesIO(r.content)).convert("RGBA").resize(
                (148, 148), Image.LANCZOS)
            ring = make_avatar_ring(raw, 148)
            ci   = ctk.CTkImage(ring, ring, size=(182, 182))
            self._put("pic", ci)
        except Exception:
            pass

    def _render_posts(self, medias):
        for w in self._posts_sc.winfo_children(): w.destroy()
        if not medias:
            ctk.CTkLabel(self._posts_sc,
                         text="No posts available (private?)",
                         font=ctk.CTkFont(size=14),
                         text_color=C["txt4"]).pack(pady=50)
            return
        tl = sum(getattr(m,"like_count",0) or 0 for m in medias)
        tc = sum(getattr(m,"comment_count",0) or 0 for m in medias)
        self._posts_stat.set(
            f"{len(medias)} posts  │  {fmt_num(tl)} total likes  │  {fmt_num(tc)} total comments")
        for i, m in enumerate(medias):
            self._post_card(m, i)

    def _post_card(self, m, idx):
        card = ctk.CTkFrame(self._posts_sc, fg_color=C["card"], corner_radius=10,
                            border_color=C["border"], border_width=1)
        card.pack(fill="x", pady=3, padx=2)

        hdr = ctk.CTkFrame(card, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(9,3))
        ta     = getattr(m,"taken_at",None)
        ds     = ta.strftime("%Y-%m-%d  %H:%M") if ta else "Unknown"
        mt     = {1:"📷",2:"🎬",8:"📚"}.get(getattr(m,"media_type",1),"📷")
        likes  = getattr(m,"like_count",0) or 0
        cmts   = getattr(m,"comment_count",0) or 0

        ctk.CTkLabel(hdr, text=f"#{idx+1}",
                     font=ctk.CTkFont(size=11,weight="bold"),
                     text_color=C["blue"], width=32).pack(side="left")
        ctk.CTkLabel(hdr, text=mt,
                     font=ctk.CTkFont(size=11),
                     text_color=C["cyan"]).pack(side="left", padx=(0,8))
        ctk.CTkLabel(hdr, text=f"📅  {ds}",
                     font=ctk.CTkFont(size=11),
                     text_color=C["txt2"]).pack(side="left")

        for txt, col in ((f"❤  {fmt_num(likes)}",C["red"]),
                          (f"💬  {fmt_num(cmts)}",C["txt3"])):
            ctk.CTkLabel(hdr, text=txt,
                         font=ctk.CTkFont(size=11),
                         text_color=col).pack(side="right", padx=5)

        cap = str(getattr(m,"caption_text","") or "")
        if cap:
            cf = ctk.CTkFrame(card, fg_color=C["bg3"], corner_radius=6)
            cf.pack(fill="x", padx=12, pady=(0,4))
            ctk.CTkLabel(cf, text=cap[:320]+("…" if len(cap)>320 else ""),
                         font=ctk.CTkFont(size=11), text_color=C["txt"],
                         anchor="w", justify="left", wraplength=1050).pack(
                             padx=10, pady=5, anchor="w")
        tags = re.findall(r"#\w+", cap)
        if tags:
            ctk.CTkLabel(card, text="  ".join(tags[:20]),
                         font=ctk.CTkFont(size=10),
                         text_color=C["purple"],
                         wraplength=1050, justify="left").pack(
                             anchor="w", padx=12, pady=(0,4))
        loc = getattr(m,"location",None)
        if loc:
            parts = [str(getattr(loc,a,"") or "") for a in ("name","city","country_code")]
            ls = "  │  ".join(p for p in parts if p)
            if ls:
                ctk.CTkLabel(card, text=f"📍  {ls}",
                             font=ctk.CTkFont(size=10),
                             text_color=C["green"]).pack(
                                 anchor="w", padx=12, pady=(0,6))

    def _render_locations(self, medias):
        tagged = []
        for m in medias:
            lo = getattr(m,"location",None)
            if lo:
                tagged.append((str(getattr(lo,"name","") or ""),
                               str(getattr(lo,"city","") or ""),
                               str(getattr(lo,"country_code","") or ""),
                               getattr(m,"taken_at",None)))
        if not tagged:
            _tb_set(self._loc_tb,"No location tags in last 50 posts."); return
        names   = Counter(l[0] for l in tagged if l[0])
        cities  = Counter(l[1] for l in tagged if l[1])
        cntries = Counter(l[2] for l in tagged if l[2])
        out = f"📍  LOCATION INTELLIGENCE  —  {len(tagged)} tagged posts\n"+"═"*66+"\n\n"
        def section(title, counter, n=20):
            nonlocal out
            out += f"{title}\n"+"─"*50+"\n"
            for nm,cnt in counter.most_common(n):
                bar = "█"*min(cnt*3,40)
                out += f"  {bar:<40}  {cnt:3d}×   {nm}\n"
            out += "\n"
        section("🏛  VENUES / PLACES",  names,   20)
        section("🌆  CITIES",           cities,  15)
        section("🌍  COUNTRIES",        cntries, 10)
        out += "📋  ALL TAGGED POSTS:\n"+"─"*50+"\n"
        for nm,ci,cc,ta in tagged:
            ds    = ta.strftime("%Y-%m-%d") if ta else "??"
            parts = "  │  ".join(p for p in [nm,ci,cc] if p)
            out  += f"  [{ds}]   {parts}\n"
        _tb_set(self._loc_tb, out)

    def _render_bio_intel(self, bio, medias):
        lines    = []
        cap_emails, cap_phones = set(), set()
        for m in medias:
            cap = str(getattr(m,"caption_text","") or "")
            cap_emails.update(EMAIL_RE.findall(cap))
            cap_phones.update(PHONE_RE.findall(cap))

        bio_emails = EMAIL_RE.findall(bio)
        bio_phones = PHONE_RE.findall(bio)
        bio_urls   = re.findall(r"https?://\S+", bio)
        bio_ats    = re.findall(r"@[\w.]+", bio)

        if bio_emails: lines.append(f"📧  Emails in bio:    {', '.join(bio_emails)}")
        if bio_phones: lines.append(f"📱  Phones in bio:    {', '.join(bio_phones)}")
        if bio_urls:   lines.append(f"🔗  URLs in bio:      {', '.join(bio_urls[:5])}")
        if bio_ats:    lines.append(f"👤  Mentions in bio:  {', '.join(bio_ats[:8])}")
        if cap_emails: lines.append(f"\n📨  Emails in captions:\n  "+"\n  ".join(sorted(cap_emails)))
        if cap_phones: lines.append(f"\n📱  Phones in captions:\n  "+"\n  ".join(sorted(cap_phones)))

        _tb_set(self._emails_tb, "\n".join(sorted(cap_emails)) if cap_emails else "None found.")
        _tb_set(self._phones_tb, "\n".join(sorted(cap_phones)) if cap_phones else "None found.")
        _tb_set(self._bio_intel_tb, "\n".join(lines) if lines else "No contact intel found.")

    def _extract_contact_data(self, user, medias):
        cap_emails, cap_phones = set(), set()
        for m in medias:
            cap = str(getattr(m,"caption_text","") or "")
            cap_emails.update(EMAIL_RE.findall(cap))
            cap_phones.update(PHONE_RE.findall(cap))
        return {
            "email":   str(getattr(user,"business_email","") or getattr(user,"public_email","") or ""),
            "phone":   str(getattr(user,"business_phone_number","") or getattr(user,"contact_phone_number","") or ""),
            "caption_emails": sorted(cap_emails),
            "caption_phones": sorted(cap_phones),
        }

    def _compute_activity(self, medias) -> dict:
        hours=[0]*24; days=[0]*7; dates=[]
        for m in medias:
            ta = getattr(m,"taken_at",None)
            if ta: hours[ta.hour]+=1; days[ta.weekday()]+=1; dates.append(ta)
        span = max((max(dates)-min(dates)).days,1) if len(dates)>=2 else 1
        avg  = round(len(medias)/span,3)
        ph   = hours.index(max(hours)) if any(hours) else 0
        pd   = days.index(max(days))   if any(days) else 0
        n    = max(len(medias),1)
        tl   = sum(getattr(m,"like_count",0) or 0 for m in medias)
        tc   = sum(getattr(m,"comment_count",0) or 0 for m in medias)
        tag_c= Counter()
        for m in medias:
            cap = str(getattr(m,"caption_text","") or "")
            tag_c.update(t.lower() for t in re.findall(r"#\w+",cap))
        return {
            "hours": hours, "days": days,
            "avg_posts_per_day": avg,
            "peak_hour":  f"{ph:02d}:00 – {ph:02d}:59",
            "peak_day":   DAY_NAMES[pd],
            "avg_likes":  tl//n, "avg_comments": tc//n,
            "last_post":  max(dates).strftime("%Y-%m-%d %H:%M") if dates else "—",
            "top_hashtags": [t for t,_ in tag_c.most_common(20)],
            "tag_counter":  dict(tag_c.most_common(20)),
        }

    def _apply_activity_ui(self, act, medias):
        self._vavgposts.set(f"{act['avg_posts_per_day']:.2f} / day")
        self._vpeak_hr.set(act["peak_hour"])
        self._vpeak_day.set(act["peak_day"])
        self._vavg_likes.set(fmt_num(act["avg_likes"]))
        self._vavg_cmts.set(fmt_num(act["avg_comments"]))
        if act["last_post"] != "—": self._vlast_post.set(act["last_post"])
        tags = act["top_hashtags"]
        _tb_set(self._tags_tb,
                "\n".join(f"{t}  ({act['tag_counter'].get(t,0)}×)" for t in tags[:12])
                if tags else "No hashtags found.")
        threading.Thread(target=self._gen_charts,
                         args=(act["hours"], act["days"],
                               Counter(act["tag_counter"]), medias),
                         daemon=True).start()

    def _gen_charts(self, hours, days, tag_counter, medias):
        try:
            charts = {}
            charts["hours"] = ChartEngine.hours_chart(hours)
            charts["days"]  = ChartEngine.days_chart(days)
            if tag_counter:
                charts["tags"]  = ChartEngine.hashtags_chart(tag_counter)
            if medias:
                charts["eng"]   = ChartEngine.engagement_chart(medias)
            for key, img in charts.items():
                if img:
                    p = Path(f"_chart_{key}.png")
                    img.save(str(p))
                    self._chart_paths[key] = str(p)
            self._put("charts", charts)
        except Exception:
            print(traceback.format_exc())

    # ══════════════════════════════════════════════════════════════════════════
    #  Downloads
    # ══════════════════════════════════════════════════════════════════════════

    def _dl_start(self):
        if not self._data:
            Toast.show(self, "Run analysis first", "warn"); return
        u = self._data.get("username","unknown")
        self._dl_btn.configure(state="disabled", text="⏳  Downloading …")
        self._dl_cancel.configure(state="normal")
        self._dl_mgr = DownloadManager(
            status_cb   = lambda m: self._put("dl_log", m),
            progress_cb = lambda d,t: self._put("dl_progress",(d,t)))

        def _run():
            if self._dl_pic_var.get():
                pic_url = str(self._data.get("profile",{}).get("profile_pic_url_hd","") or
                              self._data.get("profile",{}).get("profile_pic_url","") or "")
                if pic_url and pic_url != "None":
                    dest = self._dl_mgr.download_profile_pic(pic_url, u)
                    self._put("dl_log", f"Profile pic → {dest or 'failed'}")
            stats = {"total":0,"downloaded":0,"failed":0}
            if self._dl_post_var.get() and self._medias_cache:
                self._put("dl_log", f"Downloading {len(self._medias_cache)} posts …")
                stats = self._dl_mgr.download_medias(self._medias_cache, u, self._cl)
            self._put("dl_done", stats)

        threading.Thread(target=_run, daemon=True).start()

    def _dl_cancel(self):
        if self._dl_mgr: self._dl_mgr.cancel()
        self._dl_btn.configure(state="normal", text="⬇  Start Download")
        self._dl_cancel.configure(state="disabled")
        self._put("dl_log","⛔  Cancelled")

    # ══════════════════════════════════════════════════════════════════════════
    #  Exports
    # ══════════════════════════════════════════════════════════════════════════

    def _guard(self):
        if not self._data:
            Toast.show(self, "Run analysis first", "warn"); return False
        return True

    def _log_export(self, fn):
        if fn:
            self._put("export_log", f"✅  {fn}")
            self._typewriter(f"✅  Saved: {fn.name}")
            Toast.show(self, f"Saved: {fn.name}", "ok")
        else:
            self._put("export_log","❌  Export failed — check dependencies")
            Toast.show(self,"Export failed — check dependencies","err")

    def _save_json(self):
        if not self._guard(): return
        self._log_export(Exporter.to_json(self._data, self._data.get("username","x")))

    def _copy_json(self):
        self.clipboard_clear()
        self.clipboard_append(json.dumps(self._data, indent=2, default=str))
        Toast.show(self, "Copied to clipboard", "info")

    def _export_excel(self):
        if not self._guard(): return
        if not OPENPYXL_OK:
            Toast.show(self,"pip install openpyxl","warn"); return
        self._log_export(Exporter.to_excel(self._data, self._data.get("username","x")))

    def _export_html(self):
        if not self._guard(): return
        self._log_export(Exporter.to_html(self._data, self._data.get("username","x")))

    def _export_pdf(self):
        if not self._guard(): return
        if not REPORTLAB_OK:
            Toast.show(self,"pip install reportlab","warn"); return
        self._log_export(Exporter.to_pdf(self._data, self._data.get("username","x"),
                                         self._chart_paths or None))

    # ══════════════════════════════════════════════════════════════════════════
    #  Serialisation
    # ══════════════════════════════════════════════════════════════════════════

    def _ser(self, obj):
        if isinstance(obj, dict):     return {k:self._ser(v) for k,v in obj.items()}
        if isinstance(obj, list):     return [self._ser(v) for v in obj]
        if isinstance(obj, datetime): return obj.isoformat()
        try: json.dumps(obj); return obj
        except: return str(obj)

    def _ser_media(self, m):
        loc = getattr(m,"location",None)
        return {
            "id":       str(getattr(m,"pk","")),
            "taken_at": (getattr(m,"taken_at",None) or datetime.now()).isoformat(),
            "caption":  str(getattr(m,"caption_text","") or "")[:400],
            "likes":    getattr(m,"like_count",0) or 0,
            "comments": getattr(m,"comment_count",0) or 0,
            "type":     getattr(m,"media_type",1),
            "location": {
                "name":    str(getattr(loc,"name","") or ""),
                "city":    str(getattr(loc,"city","") or ""),
                "country": str(getattr(loc,"country_code","") or ""),
                "lat":     getattr(loc,"lat",None),
                "lng":     getattr(loc,"lng",None),
            } if loc else None,
        }


# ─────────────────────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if not INSTAGRAPI_OK:
        print("\n❌  Missing: instagrapi\n"
              "   pip install customtkinter instagrapi matplotlib pillow requests\n"
              "   Optional: pip install openpyxl reportlab tkinterdnd2 browser-cookie3\n")
        raise SystemExit(1)
    App().mainloop()
